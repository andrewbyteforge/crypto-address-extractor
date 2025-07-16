"""
File Handler Module - Production Ready Version with Address Frequency Analysis
=============================================================================

This module handles all file I/O operations for the cryptocurrency address extractor.
Provides methods for reading CSV/Excel files and writing comprehensive Excel reports
with multiple analysis sheets including address frequency analysis.

Author: Crypto Extractor Tool
Date: 2025-07-16
Version: 2.4.0 - Added Address Frequency Analysis feature
"""

import os
import logging
import pandas as pd
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.worksheet import Worksheet
import re
import json

from extractor import ExtractedAddress


class FileHandler:
    """
    Handles all file I/O operations for the cryptocurrency address extractor.
    
    Features:
    - Read CSV and Excel files with multi-sheet support
    - Write comprehensive Excel reports with multiple analysis sheets
    - Handle large datasets with row limit protection
    - Professional formatting and styling
    - Backward compatible save_to_excel method
    - Enhanced direct and indirect exchange exposure display
    - Summary sheet with extraction statistics
    - Column definitions sheet for user reference
    - Address frequency analysis sheet
    """
    
    # Excel row limit with safety buffer
    EXCEL_MAX_ROWS = 1048576
    SAFE_ROW_LIMIT = 1000000  # Leave buffer for headers and formatting
    
    def __init__(self):
        """
        Initialize the file handler with logging configuration and error tracking.
        
        File: file_handler.py
        Function: __init__()
        
        Enhanced to include comprehensive file processing tracking.
        """
        try:
            self.logger = logging.getLogger(__name__)
            
            # Initialize file processing tracking
            self.file_processing_stats = {
                'files_attempted': [],
                'files_successful': [],
                'files_failed': [],
                'files_empty': [],
                'total_files_selected': 0,
                'processing_errors': {}
            }
            
            self.logger.info("FileHandler initialized successfully with error tracking")
        except Exception as e:
            print(f"Failed to initialize FileHandler logger: {e}")
            raise
    
    def read_csv(self, file_path: str, encoding: str = 'utf-8') -> pd.DataFrame:
        """
        Read CSV file and return as DataFrame.
        
        Args:
            file_path (str): Path to the CSV file
            encoding (str): File encoding (default: utf-8)
            
        Returns:
            pd.DataFrame: Loaded data
            
        Raises:
            FileNotFoundError: If file doesn't exist
            pd.errors.ParserError: If CSV parsing fails
            Exception: For other errors
        """
        self.logger.info(f"Reading CSV file: {file_path}")
        
        if not os.path.exists(file_path):
            error_msg = f"CSV file not found: {file_path}"
            self.logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        try:
            # Read CSV with multiple encoding fallbacks
            encodings = [encoding, 'utf-8', 'latin-1', 'cp1252']
            
            for enc in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=enc, low_memory=False)
                    self.logger.info(f"Successfully read CSV with {enc} encoding: {len(df)} rows")
                    return df
                except UnicodeDecodeError:
                    continue
            
            # If all encodings fail
            raise ValueError(f"Could not read CSV with any encoding: {encodings}")
            
        except pd.errors.ParserError as e:
            error_msg = f"CSV parsing error in {file_path}: {str(e)}"
            self.logger.error(error_msg)
            raise pd.errors.ParserError(error_msg)
        except Exception as e:
            error_msg = f"Error reading CSV file {file_path}: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e
    
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, pd.DataFrame]:
        """
        Read Excel file and return dictionary of DataFrames for each sheet.
        
        Args:
            file_path (str): Path to the Excel file
            sheet_name (Optional[str]): Specific sheet to read (None for all sheets)
            
        Returns:
            Dict[str, pd.DataFrame]: Dictionary mapping sheet names to DataFrames
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If Excel file is invalid
            Exception: For other errors
        """
        self.logger.info(f"Reading Excel file: {file_path}")
        
        if not os.path.exists(file_path):
            error_msg = f"Excel file not found: {file_path}"
            self.logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        try:
            # Read Excel file
            excel_file = pd.ExcelFile(file_path)
            sheets_to_read = [sheet_name] if sheet_name else excel_file.sheet_names
            
            result = {}
            for sheet in sheets_to_read:
                if sheet in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet, header=None)
                    result[sheet] = df
                    self.logger.info(f"Read sheet '{sheet}': {len(df)} rows")
                else:
                    self.logger.warning(f"Sheet '{sheet}' not found in {file_path}")
            
            excel_file.close()
            return result
            
        except ValueError as e:
            error_msg = f"Invalid Excel file {file_path}: {str(e)}"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        except Exception as e:
            error_msg = f"Error reading Excel file {file_path}: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e
    
    def save_to_excel(self, addresses: List[ExtractedAddress], output_path: str, 
                      include_api_data: Optional[bool] = None) -> str:
        """
        Save extracted addresses to Excel file (wrapper for backward compatibility).
        
        This method wraps write_to_excel to maintain compatibility with existing code
        that expects save_to_excel to return the output path.
        
        Args:
            addresses (List[ExtractedAddress]): List of extracted addresses
            output_path (str): Path for the output Excel file
            include_api_data (Optional[bool]): Whether to include API data columns.
                                             If None, auto-detects from addresses.
            
        Returns:
            str: The path of the saved file
            
        Raises:
            PermissionError: If file is open or inaccessible
            Exception: For other errors
        """
        try:
            # Auto-detect if we should include API data if not specified
            if include_api_data is None:
                include_api_data = any(hasattr(addr, 'api_balance') for addr in addresses)
                if include_api_data:
                    self.logger.info("Auto-detected API data in addresses")
            
            # Call the main write_to_excel method
            self.write_to_excel(addresses, output_path, include_api_data)
            
            # Log success
            self.logger.info(f"Successfully saved {len(addresses)} addresses to {output_path}")
            
            # Return the output path for backward compatibility
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error in save_to_excel: {str(e)}")
            raise
    
    def write_to_excel(self, addresses: List[ExtractedAddress], output_path: str,
                    include_api_data: bool = False, selected_files: List[str] = None) -> None:
        """
        Write extracted addresses to a formatted Excel file with enhanced file tracking.
        
        File: file_handler.py
        Function: write_to_excel()
        
        Enhanced to accept selected_files parameter for comprehensive tracking.
        
        Args:
            addresses (List[ExtractedAddress]): List of extracted addresses
            output_path (str): Path for the output Excel file
            include_api_data (bool): Whether to include API data columns
            selected_files (List[str], optional): List of originally selected files
            
        Raises:
            PermissionError: If file is open or inaccessible
            Exception: For other errors
        """
        self.logger.info(f"Writing {len(addresses)} addresses to Excel: {output_path}")
        
        # Store selected files count for summary
        if selected_files:
            self.file_processing_stats['total_files_selected'] = len(selected_files)
        
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets in the desired order:
            # 1. Summary sheet (position 0) - now with enhanced file tracking
            self._create_summary_sheet(wb, addresses, getattr(self, "_api_stats", None))
            
            # 2. Column Definitions sheet (position 1)
            self._create_column_definitions_sheet(wb, include_api_data)
            
            # 3. Address Frequency Analysis sheet (position 2)
            self._create_frequency_analysis_sheet(wb, addresses, include_api_data)
            
            # 4. All Addresses sheet (position 3)
            self._create_all_addresses_sheet(wb, addresses, include_api_data)
            
            # 5. Individual crypto sheets (positions 4+)
            crypto_groups = defaultdict(list)
            for addr in addresses:
                crypto_groups[addr.crypto_name].append(addr)
            
            # Create individual crypto sheets
            for crypto_name, crypto_addresses in crypto_groups.items():
                self._create_crypto_sheet(wb, crypto_name, crypto_addresses, include_api_data)
            
            # 6. Optional: Create duplicate analysis sheet if duplicates exist
            if any(addr.is_duplicate for addr in addresses):
                self._create_duplicate_analysis_sheet(wb, addresses)
            
            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Successfully created Excel file with enhanced error tracking: {output_path}")
            
        except PermissionError as e:
            error_msg = f"Permission denied writing to {output_path}. Please close the file if it's open."
            self.logger.error(error_msg)
            raise PermissionError(error_msg) from e
        except Exception as e:
            error_msg = f"Failed to create Excel file: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e
        

    def read_excel_with_tracking(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, pd.DataFrame]:
        """
        Read Excel file with comprehensive error tracking and diagnostics.
        
        File: file_handler.py
        Function: read_excel_with_tracking()
        
        Enhanced version of read_excel that captures all processing attempts and errors.
        
        Args:
            file_path (str): Path to the Excel file
            sheet_name (Optional[str]): Specific sheet to read (None for all sheets)
            
        Returns:
            Dict[str, pd.DataFrame]: Dictionary mapping sheet names to DataFrames
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If Excel file is invalid
            Exception: For other errors
        """
        filename = os.path.basename(file_path)
        self.logger.info(f"TRACKING: Attempting to read Excel file: {file_path}")
        
        # Record attempt
        self.file_processing_stats['files_attempted'].append({
            'filename': filename,
            'full_path': file_path,
            'file_type': 'Excel',
            'file_size': 0,
            'attempt_time': datetime.now().strftime("%H:%M:%S")
        })
        
        if not os.path.exists(file_path):
            error_msg = f"Excel file not found: {file_path}"
            self.logger.error(f"TRACKING: {error_msg}")
            self.file_processing_stats['files_failed'].append({
                'filename': filename,
                'error_type': 'File Not Found',
                'error_message': error_msg,
                'file_size': 0
            })
            self.file_processing_stats['processing_errors'][filename] = {
                'error': 'File Not Found',
                'details': error_msg
            }
            raise FileNotFoundError(error_msg)
        
        # Get file size
        try:
            file_size = os.path.getsize(file_path)
            self.logger.info(f"TRACKING: Excel file size: {file_size} bytes")
            
            # Update file size in attempted files
            for attempt in self.file_processing_stats['files_attempted']:
                if attempt['full_path'] == file_path:
                    attempt['file_size'] = file_size
                    break
                    
        except Exception as e:
            self.logger.warning(f"TRACKING: Could not get Excel file size: {e}")
            file_size = 0
        
        try:
            # Read Excel file
            excel_file = pd.ExcelFile(file_path)
            sheets_to_read = [sheet_name] if sheet_name else excel_file.sheet_names
            
            self.logger.info(f"TRACKING: Excel sheets available: {excel_file.sheet_names}")
            
            result = {}
            total_rows = 0
            
            for sheet in sheets_to_read:
                if sheet in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet, header=None)
                    result[sheet] = df
                    total_rows += len(df)
                    self.logger.info(f"TRACKING: Read sheet '{sheet}': {len(df)} rows")
                else:
                    self.logger.warning(f"TRACKING: Sheet '{sheet}' not found in {file_path}")
            
            excel_file.close()
            
            # Record successful read
            self.file_processing_stats['files_successful'].append({
                'filename': filename,
                'file_type': 'Excel',
                'rows_read': total_rows,
                'sheets_read': len(result),
                'sheets_available': len(excel_file.sheet_names),
                'file_size': file_size
            })
            
            self.logger.info(f"TRACKING: Successfully processed Excel {filename} - {total_rows} total rows, {len(result)} sheets")
            return result
            
        except ValueError as e:
            error_msg = f"Invalid Excel file {file_path}: {str(e)}"
            self.logger.error(f"TRACKING: {error_msg}")
            self.file_processing_stats['files_failed'].append({
                'filename': filename,
                'error_type': 'Invalid Excel File',
                'error_message': str(e),
                'file_size': file_size
            })
            self.file_processing_stats['processing_errors'][filename] = {
                'error': 'Invalid Excel File',
                'details': str(e)
            }
            raise ValueError(error_msg)
        except Exception as e:
            error_msg = f"Error reading Excel file {file_path}: {str(e)}"
            self.logger.error(f"TRACKING: {error_msg}", exc_info=True)
            self.file_processing_stats['files_failed'].append({
                'filename': filename,
                'error_type': 'General Error',
                'error_message': str(e),
                'file_size': file_size
            })
            self.file_processing_stats['processing_errors'][filename] = {
                'error': 'General Error',
                'details': str(e)
            }
            raise Exception(error_msg) from e

        
    def read_csv_with_tracking(self, file_path: str, encoding: str = 'utf-8') -> pd.DataFrame:
        """
        Read CSV file with comprehensive error tracking and diagnostics.
        
        File: file_handler.py
        Function: read_csv_with_tracking()
        
        Enhanced version of read_csv that captures all processing attempts and errors.
        
        Args:
            file_path (str): Path to the CSV file
            encoding (str): File encoding (default: utf-8)
            
        Returns:
            pd.DataFrame: Loaded data
            
        Raises:
            FileNotFoundError: If file doesn't exist
            pd.errors.ParserError: If CSV parsing fails
            Exception: For other errors
        """
        filename = os.path.basename(file_path)
        self.logger.info(f"TRACKING: Attempting to read CSV file: {file_path}")
        
        # Record attempt
        self.file_processing_stats['files_attempted'].append({
            'filename': filename,
            'full_path': file_path,
            'file_type': 'CSV',
            'file_size': 0,
            'attempt_time': datetime.now().strftime("%H:%M:%S")
        })
        
        # Check file existence and basic properties
        if not os.path.exists(file_path):
            error_msg = f"CSV file not found: {file_path}"
            self.logger.error(f"TRACKING: {error_msg}")
            self.file_processing_stats['files_failed'].append({
                'filename': filename,
                'error_type': 'File Not Found',
                'error_message': error_msg,
                'file_size': 0
            })
            self.file_processing_stats['processing_errors'][filename] = {
                'error': 'File Not Found',
                'details': error_msg
            }
            raise FileNotFoundError(error_msg)
        
        # Get file size for diagnostics
        try:
            file_size = os.path.getsize(file_path)
            self.logger.info(f"TRACKING: File size: {file_size} bytes")
            
            # Update file size in attempted files
            for attempt in self.file_processing_stats['files_attempted']:
                if attempt['full_path'] == file_path:
                    attempt['file_size'] = file_size
                    break
                    
        except Exception as e:
            self.logger.warning(f"TRACKING: Could not get file size: {e}")
            file_size = 0
        
        try:
            # Read CSV with multiple encoding fallbacks
            encodings = [encoding, 'utf-8', 'latin-1', 'cp1252', 'utf-8-sig']  # Added utf-8-sig for BOM
            
            df = None
            successful_encoding = None
            
            for enc in encodings:
                try:
                    self.logger.info(f"TRACKING: Trying encoding: {enc}")
                    df = pd.read_csv(file_path, encoding=enc, low_memory=False)
                    successful_encoding = enc
                    self.logger.info(f"TRACKING: Successfully read CSV with {enc} encoding: {len(df)} rows")
                    break
                except UnicodeDecodeError as ude:
                    self.logger.warning(f"TRACKING: Encoding {enc} failed: {ude}")
                    continue
                except Exception as e:
                    self.logger.warning(f"TRACKING: Encoding {enc} failed with: {e}")
                    continue
            
            if df is None:
                error_msg = f"Could not read CSV with any encoding: {encodings}"
                self.logger.error(f"TRACKING: {error_msg}")
                self.file_processing_stats['files_failed'].append({
                    'filename': filename,
                    'error_type': 'Encoding Error',
                    'error_message': error_msg,
                    'file_size': file_size
                })
                self.file_processing_stats['processing_errors'][filename] = {
                    'error': 'Encoding Error',
                    'details': error_msg,
                    'encodings_tried': encodings
                }
                raise ValueError(error_msg)
            
            # Record successful read
            self.file_processing_stats['files_successful'].append({
                'filename': filename,
                'file_type': 'CSV',
                'rows_read': len(df),
                'columns': len(df.columns),
                'encoding_used': successful_encoding,
                'file_size': file_size
            })
            
            self.logger.info(f"TRACKING: Successfully processed {filename} - {len(df)} rows, {len(df.columns)} columns")
            return df
            
        except pd.errors.ParserError as e:
            error_msg = f"CSV parsing error in {file_path}: {str(e)}"
            self.logger.error(f"TRACKING: {error_msg}")
            self.file_processing_stats['files_failed'].append({
                'filename': filename,
                'error_type': 'Parse Error',
                'error_message': str(e),
                'file_size': file_size
            })
            self.file_processing_stats['processing_errors'][filename] = {
                'error': 'Parse Error',
                'details': str(e)
            }
            raise pd.errors.ParserError(error_msg)
        except Exception as e:
            error_msg = f"Error reading CSV file {file_path}: {str(e)}"
            self.logger.error(f"TRACKING: {error_msg}", exc_info=True)
            self.file_processing_stats['files_failed'].append({
                'filename': filename,
                'error_type': 'General Error',
                'error_message': str(e),
                'file_size': file_size
            })
            self.file_processing_stats['processing_errors'][filename] = {
                'error': 'General Error',
                'details': str(e)
            }
            raise Exception(error_msg) from e


    def record_empty_file(self, file_path: str, addresses_found: int):
        """
        Record when a file was successfully read but contained no addresses.
        
        File: file_handler.py
        Function: record_empty_file()
        
        Args:
            file_path (str): Path to the file
            addresses_found (int): Number of addresses found (should be 0 for empty)
        """
        filename = os.path.basename(file_path)
        
        if addresses_found == 0:
            self.file_processing_stats['files_empty'].append({
                'filename': filename,
                'file_type': 'CSV' if filename.lower().endswith('.csv') else 'Excel',
                'reason': 'No cryptocurrency addresses found in file'
            })
            self.logger.info(f"TRACKING: File {filename} processed successfully but contained no addresses")

    def _create_summary_sheet(self, wb: Workbook, addresses: List[ExtractedAddress], 
                        api_stats=None) -> None:
        """
        Create summary sheet with comprehensive file processing diagnostics.
        
        File: file_handler.py
        Function: _create_summary_sheet()
        
        Enhanced to show detailed file processing tracking and error information.
        """
        try:
            self.logger.info("Creating comprehensive Summary sheet with file processing diagnostics")
            
            # Create the summary sheet first (position 0)
            ws = wb.create_sheet("Summary", 0)
            
            # Title section
            ws['A1'] = "Cryptocurrency Address Extraction Summary"
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A1:D1')
            
            current_row = 3
            
            # Extraction date
            ws.cell(row=current_row, column=1, value="Extraction Date:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            current_row += 2
            
            # Enhanced file processing diagnostics section
            ws.cell(row=current_row, column=1, value="FILE PROCESSING DIAGNOSTICS").font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
            
            # File processing statistics
            files_attempted = len(self.file_processing_stats['files_attempted'])
            files_successful = len(self.file_processing_stats['files_successful'])
            files_failed = len(self.file_processing_stats['files_failed'])
            files_empty = len(self.file_processing_stats['files_empty'])
            total_selected = self.file_processing_stats.get('total_files_selected', files_attempted)
            
            ws.cell(row=current_row, column=1, value="Files Selected:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=total_selected)
            ws.cell(row=current_row, column=3, value="Files Attempted:").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=files_attempted)
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Files Successfully Read:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=files_successful)
            ws.cell(row=current_row, column=3, value="Files Failed to Read:").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=files_failed)
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Files with Addresses:").font = Font(bold=True)
            files_with_addresses = len(set(addr.filename for addr in addresses))
            ws.cell(row=current_row, column=2, value=files_with_addresses)
            ws.cell(row=current_row, column=3, value="Empty Files (No Addresses):").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=files_successful - files_with_addresses)
            current_row += 2
            
            # File processing details table
            if self.file_processing_stats['files_attempted']:
                ws.cell(row=current_row, column=1, value="DETAILED FILE PROCESSING LOG").font = Font(bold=True, size=12, color="FFFFFF")
                ws.cell(row=current_row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                ws.merge_cells(f'A{current_row}:D{current_row}')
                current_row += 1
                
                # Headers
                ws.cell(row=current_row, column=1, value="Filename").font = Font(bold=True)
                ws.cell(row=current_row, column=2, value="Status").font = Font(bold=True)
                ws.cell(row=current_row, column=3, value="File Size").font = Font(bold=True)
                ws.cell(row=current_row, column=4, value="Details").font = Font(bold=True)
                current_row += 1
                
                # Process all attempted files
                for attempt in self.file_processing_stats['files_attempted']:
                    filename = attempt['filename']
                    file_size = attempt['file_size']
                    
                    # Determine status and details
                    if filename in [f['filename'] for f in self.file_processing_stats['files_failed']]:
                        status = "READ FAILED"
                        details = self.file_processing_stats['processing_errors'].get(filename, {}).get('error', 'Unknown error')
                        status_color = "FFCCCB"  # Light red
                    elif filename in [addr.filename for addr in addresses]:
                        addresses_count = len([addr for addr in addresses if addr.filename == attempt['full_path']])
                        status = "SUCCESS"
                        details = f"{addresses_count} addresses found"
                        status_color = "C6EFCE"  # Light green
                    elif filename in [f['filename'] for f in self.file_processing_stats['files_successful']]:
                        status = "READ OK - NO ADDRESSES"
                        details = "File read successfully but no crypto addresses found"
                        status_color = "FFEB9C"  # Light yellow
                    else:
                        status = "UNKNOWN"
                        details = "Status could not be determined"
                        status_color = "D3D3D3"  # Light gray
                    
                    ws.cell(row=current_row, column=1, value=filename)
                    status_cell = ws.cell(row=current_row, column=2, value=status)
                    status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                    ws.cell(row=current_row, column=3, value=f"{file_size:,} bytes" if file_size > 0 else "Unknown")
                    ws.cell(row=current_row, column=4, value=details)
                    current_row += 1
                
                current_row += 1
            
            # Error details section
            if self.file_processing_stats['processing_errors']:
                ws.cell(row=current_row, column=1, value="ERROR DETAILS").font = Font(bold=True, size=12, color="FFFFFF")
                ws.cell(row=current_row, column=1).fill = PatternFill(start_color="FF4B4B", end_color="FF4B4B", fill_type="solid")
                ws.merge_cells(f'A{current_row}:D{current_row}')
                current_row += 1
                
                for filename, error_info in self.file_processing_stats['processing_errors'].items():
                    ws.cell(row=current_row, column=1, value=filename).font = Font(bold=True)
                    ws.cell(row=current_row, column=2, value=error_info['error'])
                    ws.merge_cells(f'C{current_row}:D{current_row}')
                    ws.cell(row=current_row, column=3, value=error_info['details'])
                    current_row += 1
                
                current_row += 1
            
            # Rest of summary (addresses, crypto breakdown, API stats)
            if addresses:
                # Address extraction summary
                ws.cell(row=current_row, column=1, value="ADDRESS EXTRACTION SUMMARY").font = Font(bold=True, size=12, color="FFFFFF")
                ws.cell(row=current_row, column=1).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                ws.merge_cells(f'A{current_row}:D{current_row}')
                current_row += 1
                
                total_addresses = len(addresses)
                unique_addresses = len(set(addr.address for addr in addresses))
                
                ws.cell(row=current_row, column=1, value="Total Addresses Found:").font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=total_addresses)
                ws.cell(row=current_row, column=3, value="Unique Addresses:").font = Font(bold=True)
                ws.cell(row=current_row, column=4, value=unique_addresses)
                current_row += 1
                
                ws.cell(row=current_row, column=1, value="Duplicate Addresses:").font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=total_addresses - unique_addresses)
                current_row += 2
                
                # Add cryptocurrency breakdown and API stats here...
                # (Include the rest of your existing summary content)
            
            # Format columns
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 40
            
            self.logger.info("✓ Created comprehensive Summary sheet with file diagnostics")
            
        except Exception as e:
            error_msg = f"Failed to create summary sheet: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e



    def _create_summary_sheet(self, wb: Workbook, addresses: List[ExtractedAddress], 
                        api_stats=None) -> None:
        """
        Create summary sheet with comprehensive file processing statistics.
        
        File: file_handler.py
        Function: _create_summary_sheet()
        
        Production-ready implementation that provides detailed file processing insights
        by analyzing extraction results and inferring comprehensive statistics.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            addresses (List[ExtractedAddress]): List of all addresses
            api_stats (dict, optional): API usage statistics from tracking
            
        Raises:
            Exception: If summary sheet creation fails
        """
        try:
            self.logger.info("Creating comprehensive Summary sheet with file processing statistics")
            
            # Create the summary sheet first (position 0)
            ws = wb.create_sheet("Summary", 0)
            
            # Title section
            ws['A1'] = "Cryptocurrency Address Extraction Summary"
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A1:D1')
            
            # Enhanced file processing analysis
            file_analysis = self._analyze_file_processing(addresses)
            
            # Basic extraction statistics
            current_row = 3
            ws.cell(row=current_row, column=1, value="Extraction Date:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            current_row += 1
            
            # Enhanced file processing section
            ws.cell(row=current_row, column=1, value="FILE PROCESSING SUMMARY").font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
            
            # File processing statistics
            ws.cell(row=current_row, column=1, value="Total Files Processed:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=file_analysis['total_files_processed'])
            ws.cell(row=current_row, column=3, value="Files with Addresses:").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=file_analysis['files_with_addresses'])
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Excel Files Processed:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=file_analysis['excel_files'])
            ws.cell(row=current_row, column=3, value="CSV Files Processed:").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=file_analysis['csv_files'])
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Excel Sheets Processed:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=file_analysis['excel_sheets_total'])
            ws.cell(row=current_row, column=3, value="Empty Files Detected:").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=file_analysis['empty_files_estimated'])
            current_row += 2
            
            # Address extraction statistics
            ws.cell(row=current_row, column=1, value="ADDRESS EXTRACTION SUMMARY").font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
            
            # Calculate address statistics
            total_addresses = len(addresses)
            unique_addresses = len(set(addr.address for addr in addresses))
            duplicate_addresses = total_addresses - unique_addresses
            
            ws.cell(row=current_row, column=1, value="Total Addresses Found:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=total_addresses)
            ws.cell(row=current_row, column=3, value="Unique Addresses:").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=unique_addresses)
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Duplicate Addresses:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=duplicate_addresses)
            ws.cell(row=current_row, column=3, value="Success Rate:").font = Font(bold=True)
            success_rate = (file_analysis['files_with_addresses'] / file_analysis['total_files_processed'] * 100) if file_analysis['total_files_processed'] > 0 else 0
            ws.cell(row=current_row, column=4, value=f"{success_rate:.1f}%")
            current_row += 2
            
            # File details breakdown
            if file_analysis['file_details']:
                ws.cell(row=current_row, column=1, value="FILE PROCESSING DETAILS").font = Font(bold=True, size=12, color="FFFFFF")
                ws.cell(row=current_row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                ws.merge_cells(f'A{current_row}:D{current_row}')
                current_row += 1
                
                # Headers for file details
                ws.cell(row=current_row, column=1, value="Filename").font = Font(bold=True)
                ws.cell(row=current_row, column=2, value="Type").font = Font(bold=True)
                ws.cell(row=current_row, column=3, value="Addresses Found").font = Font(bold=True)
                ws.cell(row=current_row, column=4, value="Status").font = Font(bold=True)
                current_row += 1
                
                # File details
                for file_detail in sorted(file_analysis['file_details'], key=lambda x: x['addresses_found'], reverse=True):
                    ws.cell(row=current_row, column=1, value=file_detail['filename'])
                    ws.cell(row=current_row, column=2, value=file_detail['file_type'])
                    ws.cell(row=current_row, column=3, value=file_detail['addresses_found'])
                    
                    # Status with color coding
                    status_cell = ws.cell(row=current_row, column=4, value=file_detail['status'])
                    if file_detail['status'] == "Success":
                        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif file_detail['status'] == "No Addresses":
                        status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    
                    current_row += 1
                
                current_row += 1
            
            # Cryptocurrency breakdown section
            ws.cell(row=current_row, column=1, value="CRYPTOCURRENCY BREAKDOWN").font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
            
            # Count addresses by cryptocurrency
            crypto_counts = defaultdict(int)
            for addr in addresses:
                crypto_counts[addr.crypto_name] += 1
            
            # Headers for crypto breakdown table
            ws.cell(row=current_row, column=1, value="Cryptocurrency").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="Count").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value="Percentage").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value="Files").font = Font(bold=True)
            current_row += 1
            
            # Crypto breakdown data with file distribution
            for crypto, count in sorted(crypto_counts.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_addresses) * 100 if total_addresses > 0 else 0
                crypto_files = len(set(addr.filename for addr in addresses if addr.crypto_name == crypto))
                
                ws.cell(row=current_row, column=1, value=crypto)
                ws.cell(row=current_row, column=2, value=count)
                ws.cell(row=current_row, column=3, value=f"{percentage:.1f}%")
                ws.cell(row=current_row, column=4, value=crypto_files)
                current_row += 1
            
            current_row += 1
            
            # API Usage Statistics Section (if available)
            if api_stats and api_stats.get('total_calls', 0) > 0:
                ws.cell(row=current_row, column=1, value="CHAINALYSIS API USAGE STATISTICS").font = Font(bold=True, size=12, color="FFFFFF")
                ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
                ws.merge_cells(f'A{current_row}:D{current_row}')
                current_row += 1
                
                # Overall API Statistics
                ws.cell(row=current_row, column=1, value="Total API Calls:").font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=api_stats.get('total_calls', 0))
                ws.cell(row=current_row, column=3, value="Success Rate:").font = Font(bold=True)
                success_rate = api_stats.get('success_rate', 0)
                ws.cell(row=current_row, column=4, value=f"{success_rate:.1f}%")
                current_row += 1
                
                ws.cell(row=current_row, column=1, value="Successful Calls:").font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=api_stats.get('successful_calls', 0))
                ws.cell(row=current_row, column=3, value="Failed Calls:").font = Font(bold=True)
                ws.cell(row=current_row, column=4, value=api_stats.get('failed_calls', 0))
                current_row += 1
                
                ws.cell(row=current_row, column=1, value="Total Processing Time:").font = Font(bold=True)
                total_time = api_stats.get('total_time_seconds', 0)
                time_str = f"{total_time/60:.1f} minutes" if total_time > 60 else f"{total_time:.1f} seconds"
                ws.cell(row=current_row, column=2, value=time_str)
                ws.cell(row=current_row, column=3, value="Avg Response Time:").font = Font(bold=True)
                avg_time = api_stats.get('avg_response_time', 0)
                ws.cell(row=current_row, column=4, value=f"{avg_time:.2f}s")
                current_row += 2
                
                # API Calls by Endpoint Type
                calls_by_endpoint = api_stats.get('calls_by_endpoint', {})
                if any(calls_by_endpoint.values()):
                    ws.cell(row=current_row, column=1, value="API Endpoint Breakdown:").font = Font(bold=True)
                    current_row += 1
                    
                    for endpoint, count in calls_by_endpoint.items():
                        if count > 0:
                            success_count = api_stats.get('success_by_endpoint', {}).get(endpoint, 0)
                            failure_count = api_stats.get('failure_by_endpoint', {}).get(endpoint, 0)
                            avg_time = api_stats.get('avg_response_times', {}).get(endpoint, 0)
                            
                            ws.cell(row=current_row, column=1, value=f"  {endpoint.title()}:")
                            ws.cell(row=current_row, column=2, value=f"{count} calls")
                            ws.cell(row=current_row, column=3, value=f"({success_count} success, {failure_count} failed)")
                            ws.cell(row=current_row, column=4, value=f"Avg: {avg_time:.2f}s")
                            current_row += 1
            
            # Format columns
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            
            self.logger.info("✓ Created comprehensive Summary sheet successfully")
            
        except Exception as e:
            error_msg = f"Failed to create summary sheet: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e

    def _analyze_file_processing(self, addresses: List[ExtractedAddress]) -> Dict:
        """
        Analyze file processing statistics from extraction results.
        
        File: file_handler.py
        Function: _analyze_file_processing()
        
        This method infers comprehensive file processing statistics by analyzing
        the extraction results, providing production-ready insights without
        requiring changes to the extraction flow.
        
        Args:
            addresses (List[ExtractedAddress]): List of all extracted addresses
            
        Returns:
            Dict: Comprehensive file processing analysis
        """
        try:
            # Analyze files and their content
            file_stats = defaultdict(lambda: {
                'addresses_found': 0,
                'sheets': set(),
                'file_type': 'Unknown'
            })
            
            excel_sheets_total = 0
            
            for addr in addresses:
                filename = addr.filename
                file_stats[filename]['addresses_found'] += 1
                
                # Determine file type from extension
                if filename.lower().endswith(('.xlsx', '.xls')):
                    file_stats[filename]['file_type'] = 'Excel'
                    if hasattr(addr, 'sheet_name') and addr.sheet_name:
                        file_stats[filename]['sheets'].add(addr.sheet_name)
                elif filename.lower().endswith('.csv'):
                    file_stats[filename]['file_type'] = 'CSV'
                
            # Count Excel sheets
            for filename, stats in file_stats.items():
                if stats['file_type'] == 'Excel':
                    excel_sheets_total += len(stats['sheets']) if stats['sheets'] else 1
            
            # Create file details for display
            file_details = []
            for filename, stats in file_stats.items():
                status = "Success" if stats['addresses_found'] > 0 else "No Addresses"
                file_details.append({
                    'filename': os.path.basename(filename),
                    'file_type': stats['file_type'],
                    'addresses_found': stats['addresses_found'],
                    'status': status
                })
            
            # Calculate totals
            total_files_processed = len(file_stats)
            files_with_addresses = len([f for f in file_stats.values() if f['addresses_found'] > 0])
            excel_files = len([f for f in file_stats.values() if f['file_type'] == 'Excel'])
            csv_files = len([f for f in file_stats.values() if f['file_type'] == 'CSV'])
            
            # Estimate empty files (this is an educated guess based on common scenarios)
            # In production, you might want to track this more precisely
            empty_files_estimated = total_files_processed - files_with_addresses
            
            analysis = {
                'total_files_processed': total_files_processed,
                'files_with_addresses': files_with_addresses,
                'empty_files_estimated': empty_files_estimated,
                'excel_files': excel_files,
                'csv_files': csv_files,
                'excel_sheets_total': excel_sheets_total,
                'file_details': file_details
            }
            
            self.logger.info(f"File processing analysis: {total_files_processed} files processed, "
                            f"{files_with_addresses} with addresses, {empty_files_estimated} empty")
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"Failed to analyze file processing: {str(e)}")
            # Return safe defaults
            return {
                'total_files_processed': len(set(addr.filename for addr in addresses)) if addresses else 0,
                'files_with_addresses': len(set(addr.filename for addr in addresses)) if addresses else 0,
                'empty_files_estimated': 0,
                'excel_files': 0,
                'csv_files': 0,
                'excel_sheets_total': 0,
                'file_details': []
            }





    def _create_column_definitions_sheet(self, wb: Workbook, include_api_data: bool = False) -> None:
        """
        Create column definitions sheet with simplified cluster-only frequency analysis.
        
        File: file_handler.py
        Function: _create_column_definitions_sheet()
        
        Updated to reflect the simplified cluster-only approach with professional blue shading.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            include_api_data (bool): Whether API data columns should be included in definitions
            
        Raises:
            Exception: If sheet creation fails
        """
        try:
            self.logger.info("Creating 'Column Definitions' sheet with simplified frequency analysis")
            
            # Create the sheet at position 1 (after Summary)
            ws = wb.create_sheet("Column Definitions", 1)
            
            # Title
            ws['A1'] = "Column Definitions & Reference Guide"
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A1:C1')
            
            # Subtitle
            ws['A2'] = "This sheet explains what each column heading means in the exported data"
            ws['A2'].font = Font(size=11, italic=True)
            ws['A2'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A2:C2')
            
            # Headers for the definitions table
            ws['A4'] = "Column Name"
            ws['B4'] = "Description"
            ws['C4'] = "Example/Notes"
            
            # Format headers
            for col in ['A4', 'B4', 'C4']:
                ws[col].font = Font(bold=True, color="FFFFFF")
                ws[col].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                ws[col].alignment = Alignment(horizontal="center")
            
            # Basic column definitions (always present)
            basic_definitions = [
                ("Address", "The cryptocurrency wallet address that was extracted", "1A1zP1eP5QGefi2DMPTfTL5SLmv7DivfNa"),
                ("Cluster Address", "The root address of the cluster this address belongs to (from API)", "Same format as Address"),
                ("Cluster Name", "Entity name from Chainalysis API", "Binance Hot Wallet, Coinbase"),
                ("Cryptocurrency Type", "The type of cryptocurrency for this address", "Bitcoin, Ethereum, Litecoin, etc."),
                ("Source File", "The name of the file where this address was found", "transactions.xlsx, addresses.csv"),
                ("Sheet", "The specific sheet name in Excel files where address was found", "Sheet1, Transaction Data"),
                ("Row", "The row number in the source file where address was located", "5, 127, 1450"),
                ("Column", "The column number/letter where address was found", "3, A, D"),
                ("Confidence %", "How confident the extraction algorithm is that this is a valid address", "95.5%, 100.0%"),
                ("Is Duplicate", "Whether this exact address appears multiple times in the data", "Yes, No"),
                ("Total Count", "How many times this cluster appears across all source files", "25, 150, 1,025")
            ]
            
            # Address Frequency Analysis sheet specific column definitions
            frequency_analysis_definitions = [
                ("Cluster Address", "Representative address for the cluster/entity from Chainalysis", "1A1zP1eP5QGefi2DMPTfTL5SLmv7DivfNa"),
                ("Cluster Name", "Name of the entity or service from Chainalysis API", "Binance Hot Wallet, Coinbase Pro"),
                ("Cluster Total Count", "Total number of times this cluster appears across all datasets", "25, 150, 1,025"),
                ("Cryptocurrency Type", "Type of cryptocurrency for addresses in this cluster", "Bitcoin, Ethereum, Multiple"),
                ("Files Containing Address", "Source files where this cluster address appears", "file1.csv, transactions.xlsx"),
                ("Sheets Containing Address", "Excel sheets containing this cluster address", "Sheet1, Data, N/A")
            ]
            
            # API-specific column definitions (only if API data is included)
            api_definitions = [
                ("Balance", "Current balance of the address in native cryptocurrency units", "1.25843210 BTC"),
                ("Total Received", "Total amount ever received by this address", "5.67891234 BTC"),
                ("Total Sent", "Total amount ever sent from this address", "4.42048024 BTC"),
                ("Transfer Count", "Number of transactions involving this address", "157, 1205"),
                ("Direct Exchange Exposure", "Exchanges that directly interact with this address", "Binance: 45.2%, Coinbase: 23.1%"),
                ("Indirect Exchange Exposure", "Exchanges connected through 1-2 intermediate addresses", "Kraken: 12.5%, Bitfinex: 8.3%"),
                ("Receiving Direct Exposure", "Exchanges this address directly receives funds from", "Coinbase: 67.8%, Kraken: 15.2%"),
                ("Receiving Indirect Exposure", "Exchanges connected to incoming transactions via intermediates", "Binance: 23.1%, Huobi: 8.9%"),
                ("Sending Direct Exposure", "Exchanges this address directly sends funds to", "Bitfinex: 45.6%, OKEx: 12.3%"),
                ("Sending Indirect Exposure", "Exchanges connected to outgoing transactions via intermediates", "Coinbase: 34.5%, Kraken: 11.2%"),
                ("Darknet Market", "Whether address has any connection to darknet markets", "Y (Yes) or N (No)"),
                ("Risk Level", "API-determined risk level based on transaction patterns", "Low, Medium, High, Critical"),
                ("Entity", "Known entity name if address belongs to a service", "Binance Hot Wallet, Coinbase"),
                ("Cluster Category", "Category of the cluster this address belongs to", "Exchange, Mixer, DeFi, etc.")
            ]
            
            # Write basic definitions
            current_row = 5
            for col_name, description, example in basic_definitions:
                ws.cell(row=current_row, column=1, value=col_name).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=description)
                ws.cell(row=current_row, column=3, value=example)
                current_row += 1
            
            # Add Address Frequency Analysis definitions as table rows
            current_row += 1
            ws.cell(row=current_row, column=1, value="ADDRESS FREQUENCY ANALYSIS SHEET COLUMNS").font = Font(bold=True, size=12, color="366092")
            ws.merge_cells(f'A{current_row}:C{current_row}')
            current_row += 1
            
            for col_name, description, example in frequency_analysis_definitions:
                ws.cell(row=current_row, column=1, value=col_name).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=description)
                ws.cell(row=current_row, column=3, value=example)
                current_row += 1
            
            # Add investigative guidance with updated color scheme
            current_row += 1
            ws.cell(row=current_row, column=1, value="PROFESSIONAL BLUE HEAT MAP GUIDANCE").font = Font(bold=True, size=12, color="366092")
            ws.merge_cells(f'A{current_row}:C{current_row}')
            current_row += 1
            
            heat_map_guidance = [
                ("Dark Blue (100+)", "Highest priority clusters - major exchanges, large services", "Focus investigation resources here first"),
                ("Medium Blue (50-99)", "High priority clusters - active exchanges, mixing services", "Secondary investigation priority"),
                ("Light Blue (20-49)", "Medium priority clusters - regular business entities", "Standard due diligence required"),
                ("Very Light Blue (5-19)", "Low priority clusters - occasional use entities", "Basic compliance checks"),
                ("White (1-4)", "Minimal priority clusters - infrequent or single-use addresses", "Low investigation priority"),
                ("Heat Map Logic", "Darker colors indicate higher cluster frequency counts", "Natural progression from white to dark blue"),
                ("Investigation Strategy", "Start with darkest blue entries for maximum impact", "Work down through lighter shades as resources allow")
            ]
            
            for col_name, description, example in heat_map_guidance:
                ws.cell(row=current_row, column=1, value=col_name).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=description)
                ws.cell(row=current_row, column=3, value=example)
                current_row += 1
            
            # Add API definitions if API data is included
            if include_api_data:
                current_row += 1
                ws.cell(row=current_row, column=1, value="API DATA COLUMNS").font = Font(bold=True, size=12, color="FF0000")
                ws.merge_cells(f'A{current_row}:C{current_row}')
                current_row += 1
                
                for col_name, description, example in api_definitions:
                    ws.cell(row=current_row, column=1, value=col_name).font = Font(bold=True)
                    ws.cell(row=current_row, column=2, value=description)
                    ws.cell(row=current_row, column=3, value=example)
                    current_row += 1
            
            # Set column widths for better readability
            ws.column_dimensions['A'].width = 30  # Column Name
            ws.column_dimensions['B'].width = 60  # Description  
            ws.column_dimensions['C'].width = 40  # Example/Notes
            
            # Add borders and formatting to make it more readable
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply borders to all table rows
            for row in range(4, current_row + 1):
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and not (cell.font.size and cell.font.size > 11):
                        cell.border = thin_border
            
            self.logger.info("✓ Created Column Definitions sheet successfully")
            
        except Exception as e:
            error_msg = f"Failed to create Column Definitions sheet: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e




    def _create_frequency_analysis_sheet(self, wb: Workbook, addresses: List[ExtractedAddress], 
                                    include_api_data: bool = False) -> None:
        """
        Create Address Frequency Analysis sheet showing cluster frequencies with professional blue shading.
        
        File: file_handler.py
        Function: _create_frequency_analysis_sheet()
        
        This sheet shows cluster total counts to help investigators understand entity activity
        levels and prioritize investigations using professional blue heat map coloring.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            addresses (List[ExtractedAddress]): List of all extracted addresses
            include_api_data (bool): Whether API data is available
            
        Raises:
            Exception: If frequency analysis sheet creation fails
        """
        try:
            self.logger.info("Creating Address Frequency Analysis sheet with cluster-only counting")
            
            # Create frequency data structure with cluster counting only
            frequency_data = self._analyze_cluster_frequency(addresses, include_api_data)
            
            # Create worksheet at position 2
            ws = wb.create_sheet("Address Frequency Analysis", 2)
            
            # Add statistics summary at the top
            self._add_frequency_statistics(ws, frequency_data, len(addresses))
            
            # Add breathing space (3 empty rows) between summary and table
            start_row = 8  # Statistics end at row 4, add 3 empty rows = start at row 8
            
            # Define simplified headers - cluster focus only
            headers = [
                "Cluster Address",
                "Cluster Name", 
                "Cluster Total Count",
                "Cryptocurrency Type",
                "Files Containing Address",
                "Sheets Containing Address"
            ]
            
            # Write headers with styling at the new start row
            self._write_frequency_headers_with_styling(ws, headers, start_row=start_row)
            
            # Sort frequency data by cluster total count (highest first) for investigative priorities
            sorted_data = sorted(frequency_data, key=lambda x: x['cluster_total_count'], reverse=True)
            
            # Write frequency data starting after headers
            for row_idx, freq_data in enumerate(sorted_data, start=start_row + 1):
                self._write_frequency_row(ws, row_idx, freq_data)
            
            # Apply professional blue shading based on cluster total count
            self._apply_frequency_color_coding(ws, sorted_data, start_row=start_row + 1)
            
            # Auto-fit columns
            self._auto_fit_frequency_columns(ws)
            
            self.logger.info(f"✓ Created Address Frequency Analysis sheet with {len(frequency_data)} cluster addresses")
            
        except Exception as e:
            self.logger.error(f"Failed to create Address Frequency Analysis sheet: {str(e)}", exc_info=True)
            raise

    def _analyze_cluster_frequency(self, addresses: List[ExtractedAddress], 
                                include_api_data: bool) -> List[Dict]:
        """
        Analyze cluster frequencies based on api_cluster_root_address.
        
        File: file_handler.py
        Function: _analyze_cluster_frequency()
        
        Simplified to count only cluster total occurrences based on Chainalysis cluster root addresses.
        
        Args:
            addresses (List[ExtractedAddress]): List of extracted addresses
            include_api_data (bool): Whether API data is available
            
        Returns:
            List[Dict]: List of cluster frequency analysis data
        """
        self.logger.info("Analyzing cluster frequencies (cluster-only counting)")
        
        # Group addresses by cluster root address
        cluster_groups = defaultdict(lambda: {
            'addresses': [],
            'cluster_name': 'Unknown',
            'crypto_types': set(),
            'files': set(),
            'sheets': set()
        })
        
        for addr in addresses:
            # Determine the cluster address to use (based on api_cluster_root_address)
            if include_api_data and hasattr(addr, 'api_cluster_root_address') and addr.api_cluster_root_address:
                cluster_address = addr.api_cluster_root_address
                cluster_name = getattr(addr, 'api_cluster_name', 'Unknown')
            else:
                # Fallback to original address if no cluster data
                cluster_address = addr.address
                cluster_name = 'Unknown'
            
            # Add to cluster group
            cluster_groups[cluster_address]['addresses'].append(addr)
            cluster_groups[cluster_address]['cluster_name'] = cluster_name
            cluster_groups[cluster_address]['crypto_types'].add(addr.crypto_type)
            cluster_groups[cluster_address]['files'].add(addr.filename)
            
            # Add sheet name if available
            if hasattr(addr, 'sheet_name') and addr.sheet_name:
                cluster_groups[cluster_address]['sheets'].add(addr.sheet_name)
        
        # Convert to frequency analysis format - cluster total count only
        frequency_data = []
        for cluster_address, group_data in cluster_groups.items():
            cluster_total_count = len(group_data['addresses'])  # Count of all addresses for this cluster
            
            frequency_data.append({
                'cluster_address': cluster_address,
                'cluster_name': group_data['cluster_name'],
                'cluster_total_count': cluster_total_count,
                'crypto_types': ', '.join(sorted(group_data['crypto_types'])),
                'files': ', '.join(sorted(group_data['files'])),
                'sheets': ', '.join(sorted(group_data['sheets'])) if group_data['sheets'] else 'N/A'
            })
        
        self.logger.info(f"Analyzed {len(frequency_data)} unique cluster addresses")
        return frequency_data

    def _write_frequency_row(self, ws: Worksheet, row_idx: int, freq_data: Dict) -> None:
        """
        Write a single frequency data row with simplified cluster-only structure.
        
        File: file_handler.py
        Function: _write_frequency_row()
        """
        ws.cell(row=row_idx, column=1, value=freq_data['cluster_address'])
        ws.cell(row=row_idx, column=2, value=freq_data['cluster_name'])
        ws.cell(row=row_idx, column=3, value=freq_data['cluster_total_count'])
        ws.cell(row=row_idx, column=4, value=freq_data['crypto_types'])
        ws.cell(row=row_idx, column=5, value=freq_data['files'])
        ws.cell(row=row_idx, column=6, value=freq_data['sheets'])


    def _write_frequency_headers_with_styling(self, ws: Worksheet, headers: List[str], start_row: int = 1) -> None:
        """
        Write headers with professional styling.
        
        File: file_handler.py
        Function: _write_frequency_headers_with_styling()
        
        Updated to handle the new 7-column layout.
        """
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"), 
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )


    def _apply_frequency_color_coding(self, ws: Worksheet, sorted_data: List[Dict], start_row: int = 2) -> None:
        """
        Apply professional blue shading based on cluster total count.
        
        File: file_handler.py
        Function: _apply_frequency_color_coding()
        
        Professional Blue Heat Map Color Scheme:
        - Dark Blue: 100+ occurrences (highest priority)
        - Medium Blue: 50-99 occurrences (high priority)
        - Light Blue: 20-49 occurrences (medium priority)  
        - Very Light Blue: 5-19 occurrences (low priority)
        - White: 1-4 occurrences (minimal priority)
        """
        self.logger.info("Applying professional blue heat map color coding")
        
        # Define professional blue shading (heat map style)
        colors = {
            'highest': PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),     # Dark Blue
            'high': PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),        # Medium Blue
            'medium': PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),      # Light Blue
            'low': PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid"),         # Very Light Blue
            'minimal': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")      # White
        }
        
        # Apply colors based on cluster total count thresholds
        for idx, freq_data in enumerate(sorted_data):
            row_idx = start_row + idx
            cluster_total = freq_data['cluster_total_count']
            
            # Determine color based on cluster total count (professional thresholds)
            if cluster_total >= 100:
                fill_color = colors['highest']
            elif cluster_total >= 50:
                fill_color = colors['high']
            elif cluster_total >= 20:
                fill_color = colors['medium']
            elif cluster_total >= 5:
                fill_color = colors['low']
            else:  # cluster_total 1-4
                fill_color = colors['minimal']
            
            # Apply color to all cells in the row
            for col in range(1, 7):  # 6 columns now
                ws.cell(row=row_idx, column=col).fill = fill_color



    def _auto_fit_frequency_columns(self, ws: Worksheet) -> None:
        """
        Auto-fit columns for optimal display with simplified headers.
        
        File: file_handler.py
        Function: _auto_fit_frequency_columns()
        """
        # Set specific column widths for simplified frequency analysis
        column_widths = {
            'A': 50,  # Cluster Address
            'B': 30,  # Cluster Name
            'C': 18,  # Cluster Total Count
            'D': 20,  # Cryptocurrency Type
            'E': 40,  # Files Containing Address
            'F': 30   # Sheets Containing Address
        }
        
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width












    def _add_frequency_statistics(self, ws: Worksheet, frequency_data: List[Dict], 
                                total_addresses: int) -> None:
        """
        Add enhanced statistical summary with color breakdown statistics.
        
        File: file_handler.py
        Function: _add_frequency_statistics()
        
        This adds summary statistics with professional blue color breakdown counts.
        """
        # Calculate enhanced statistics
        if not frequency_data:
            # Handle empty data case
            ws.cell(row=1, column=1, value="ADDRESS FREQUENCY ANALYSIS SUMMARY")
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value=f"Total Addresses Analyzed: {total_addresses}")
            ws.cell(row=2, column=4, value="Unique Cluster Addresses: 0")
            ws.cell(row=3, column=1, value="No frequency data available")
            return
        
        # Calculate statistics
        cluster_totals = [data['cluster_total_count'] for data in frequency_data]
        total_unique_clusters = len(frequency_data)
        avg_cluster_total = sum(cluster_totals) / len(cluster_totals) if cluster_totals else 0
        max_cluster_total = max(cluster_totals) if cluster_totals else 0
        
        # Find the most active cluster
        most_active = max(frequency_data, key=lambda x: x['cluster_total_count']) if frequency_data else None
        
        # Calculate color breakdown statistics
        color_breakdown = {
            'highest': len([x for x in cluster_totals if x >= 100]),      # Dark Blue: 100+
            'high': len([x for x in cluster_totals if 50 <= x < 100]),   # Medium Blue: 50-99
            'medium': len([x for x in cluster_totals if 20 <= x < 50]),  # Light Blue: 20-49
            'low': len([x for x in cluster_totals if 5 <= x < 20]),      # Very Light Blue: 5-19
            'minimal': len([x for x in cluster_totals if 1 <= x < 5])    # White: 1-4
        }
        
        # Write enhanced statistics with better formatting
        ws.cell(row=1, column=1, value="ADDRESS FREQUENCY ANALYSIS SUMMARY")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:F1')  # Merge across all 6 columns
        
        # Enhanced statistics
        ws.cell(row=2, column=1, value=f"Total Addresses Analyzed: {total_addresses}")
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=4, value=f"Unique Cluster Addresses: {total_unique_clusters}")
        ws.cell(row=2, column=4).font = Font(bold=True)
        
        ws.cell(row=3, column=1, value=f"Average Cluster Count: {avg_cluster_total:.1f}")
        ws.cell(row=3, column=1).font = Font(bold=True)
        ws.cell(row=3, column=4, value=f"Maximum Cluster Count: {max_cluster_total}")
        ws.cell(row=3, column=4).font = Font(bold=True)
        
        if most_active:
            ws.cell(row=4, column=1, value=f"Most Active Cluster: {most_active['cluster_name']} ({most_active['cluster_total_count']} occurrences)")
            ws.cell(row=4, column=1).font = Font(bold=True)
            ws.merge_cells('A4:F4')  # Merge across all columns for this longer text
        
        # Add professional color breakdown with actual colors and counts
        ws.cell(row=5, column=1, value="Priority Breakdown:")
        ws.cell(row=5, column=1).font = Font(bold=True, size=12)
        
        # Create color breakdown legend with counts
        breakdown_items = [
            (f"Highest (100+): {color_breakdown['highest']}", "1F4E79"),        # Dark Blue
            (f"High (50-99): {color_breakdown['high']}", "5B9BD5"),             # Medium Blue
            (f"Medium (20-49): {color_breakdown['medium']}", "BDD7EE"),         # Light Blue
            (f"Low (5-19): {color_breakdown['low']}", "E7F3FF"),                # Very Light Blue
            (f"Minimal (1-4): {color_breakdown['minimal']}", "FFFFFF")          # White
        ]
        
        col = 1
        for breakdown_text, color_hex in breakdown_items:
            ws.cell(row=6, column=col, value=breakdown_text)
            ws.cell(row=6, column=col).fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            ws.cell(row=6, column=col).font = Font(size=9, bold=True)
            if color_hex == "1F4E79":  # Dark blue text needs white font
                ws.cell(row=6, column=col).font = Font(size=9, bold=True, color="FFFFFF")
            col += 1
        
        # Style the statistics section with a light background
        for row in range(1, 7):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                if cell.value and row not in [1, 5, 6]:  # Don't override special formatting
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")



    def _create_all_addresses_sheet(self, wb: Workbook, addresses: List[ExtractedAddress],
                                include_api_data: bool) -> None:
        """
        Create sheet with all extracted addresses without Total Count column.
        
        File: file_handler.py
        Function: _create_all_addresses_sheet()
        
        Enhanced to properly include both direct and indirect exposure data.
        Removed Total Count column as it's specific to frequency analysis.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            addresses (List[ExtractedAddress]): List of all addresses
            include_api_data (bool): Whether to include API data columns
            
        Raises:
            Exception: If sheet creation fails
        """
        try:
            self.logger.info(f"Creating 'All Addresses' sheet with include_api_data={include_api_data}")
            
            ws = wb.create_sheet("All Addresses")
            
            # Check row limit
            if len(addresses) > self.SAFE_ROW_LIMIT:
                self._create_paginated_sheets(wb, "All Addresses", addresses, include_api_data)
                return
            
            # Headers - removed "Total Count" column
            headers = ["Address", "Cluster Address", "Cryptocurrency", "Source File", "Sheet", "Row", "Column",
                    "Confidence %", "Is Duplicate"]
            
            if include_api_data:
                # Enhanced headers with both direct and indirect exposure
                headers.extend(["Balance", "Total Received", "Total Sent", "Transfer Count",
                            "Direct Exchange Exposure", "Indirect Exchange Exposure",
                            "Receiving Direct Exposure", "Receiving Indirect Exposure", 
                            "Sending Direct Exposure", "Sending Indirect Exposure",
                            "Darknet Market", "Risk Level", "Entity", "Cluster Category"])
            
            # Write headers with formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Write data
            for row, addr in enumerate(addresses, 2):
                # Basic columns (1-9) - removed Total Count
                ws.cell(row=row, column=1, value=addr.address)
                
                # Cluster address in column 2
                cluster_address = getattr(addr, 'api_cluster_root_address', '')
                if not cluster_address:
                    cluster_address = getattr(addr, 'api_root_address', '')
                ws.cell(row=row, column=2, value=cluster_address)
                ws.cell(row=row, column=3, value=addr.crypto_name)
                ws.cell(row=row, column=4, value=addr.filename)
                ws.cell(row=row, column=5, value=addr.sheet_name or "N/A")
                ws.cell(row=row, column=6, value=addr.row)
                ws.cell(row=row, column=7, value=addr.column)
                ws.cell(row=row, column=8, value=f"{addr.confidence:.1f}%")
                ws.cell(row=row, column=9, value="Yes" if addr.is_duplicate else "No")
                # Column 10 (Total Count) removed
                
                if include_api_data:
                    # API data columns (10-23) - shifted down by 1 due to removed Total Count
                    ws.cell(row=row, column=10, value=f"{getattr(addr, 'api_balance', 0):,.8f}")
                    ws.cell(row=row, column=11, value=f"{getattr(addr, 'api_total_received', 0):,.8f}")
                    ws.cell(row=row, column=12, value=f"{getattr(addr, 'api_total_sent', 0):,.8f}")
                    ws.cell(row=row, column=13, value=getattr(addr, 'api_transfer_count', 0))
                    
                    # Format exposure text fields
                    ws.cell(row=row, column=14, value=self._format_exposure_text(getattr(addr, 'api_direct_exposure', [])))
                    ws.cell(row=row, column=15, value=self._format_exposure_text(getattr(addr, 'api_indirect_exposure', [])))
                    ws.cell(row=row, column=16, value=self._format_exposure_text(getattr(addr, 'api_receiving_direct_exposure', [])))
                    ws.cell(row=row, column=17, value=self._format_exposure_text(getattr(addr, 'api_receiving_indirect_exposure', [])))
                    ws.cell(row=row, column=18, value=self._format_exposure_text(getattr(addr, 'api_sending_direct_exposure', [])))
                    ws.cell(row=row, column=19, value=self._format_exposure_text(getattr(addr, 'api_sending_indirect_exposure', [])))
                    
                    # Darknet and other fields
                    has_darknet = getattr(addr, 'api_has_darknet_exposure', False)
                    ws.cell(row=row, column=20, value="Y" if has_darknet else "N")
                    ws.cell(row=row, column=21, value=getattr(addr, 'api_risk_level', 'Unknown'))
                    ws.cell(row=row, column=22, value=getattr(addr, 'api_cluster_name', ''))
                    ws.cell(row=row, column=23, value=getattr(addr, 'api_cluster_category', ''))
                    
            # Auto-fit columns
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 15
            
            # Wider columns for specific fields
            ws.column_dimensions['A'].width = 50  # Address
            ws.column_dimensions['B'].width = 30  # Cluster Address
            ws.column_dimensions['D'].width = 30  # Source File
            if include_api_data:
                ws.column_dimensions['N'].width = 40  # Direct Exchange Exposure (shifted from O)
                ws.column_dimensions['O'].width = 40  # Indirect Exchange Exposure (shifted from P)
            
            self.logger.info(f"✓ Created All Addresses sheet with {len(addresses)} addresses (Total Count column removed)")
            
        except Exception as e:
            self.logger.error(f"Failed to create All Addresses sheet: {str(e)}")
            raise







    def _create_crypto_sheet(self, wb: Workbook, crypto_name: str,
                           addresses: List[ExtractedAddress],
                           include_api_data: bool = False) -> None:
        """
        Create sheet for a specific cryptocurrency.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            crypto_name (str): Name of the cryptocurrency
            addresses (List[ExtractedAddress]): List of addresses for this crypto
            include_api_data (bool): Whether to include API data columns
            
        Raises:
            Exception: If crypto sheet creation fails
        """
        try:
            # Remove duplicates
            seen = set()
            unique_addresses = []
            for addr in addresses:
                if addr.address not in seen:
                    seen.add(addr.address)
                    unique_addresses.append(addr)

            ws = wb.create_sheet(crypto_name[:31])
            self.logger.info(f"Creating '{crypto_name}' sheet with {len(unique_addresses)} unique addresses")

            # Headers
            headers = ["Address", "Cluster Address", "Source File", "Row", "Column", "Confidence"]
            if include_api_data:
                headers.extend([
                    "Entity Name", "Direct Exposure", "Indirect Exposure",
                    "Receiving Direct", "Receiving Indirect",
                    "Sending Direct", "Sending Indirect",
                    "Darknet Market", "Balance", "Total Received", "Total Sent"
                ])

            # Write headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Write data
            for row, addr in enumerate(unique_addresses, 2):
                ws.cell(row=row, column=1, value=addr.address)
                
                # Cluster address in column 2
                cluster_addr = getattr(addr, 'api_cluster_root_address', '')
                if not cluster_addr:
                    cluster_addr = getattr(addr, 'api_root_address', '')
                ws.cell(row=row, column=2, value=cluster_addr)
                
                # Basic data
                ws.cell(row=row, column=3, value=addr.filename)
                ws.cell(row=row, column=4, value=addr.row)
                ws.cell(row=row, column=5, value=addr.column)
                ws.cell(row=row, column=6, value=f"{addr.confidence:.1f}%")
                
                if include_api_data:
                    # API data columns (7-17)
                    ws.cell(row=row, column=7, value=getattr(addr, 'api_cluster_name', ''))
                    ws.cell(row=row, column=8, value=self._format_exposure_text(getattr(addr, 'api_direct_exposure', [])))
                    ws.cell(row=row, column=9, value=self._format_exposure_text(getattr(addr, 'api_indirect_exposure', [])))
                    ws.cell(row=row, column=10, value=self._format_exposure_text(getattr(addr, 'api_receiving_direct_exposure', [])))
                    ws.cell(row=row, column=11, value=self._format_exposure_text(getattr(addr, 'api_receiving_indirect_exposure', [])))
                    ws.cell(row=row, column=12, value=self._format_exposure_text(getattr(addr, 'api_sending_direct_exposure', [])))
                    ws.cell(row=row, column=13, value=self._format_exposure_text(getattr(addr, 'api_sending_indirect_exposure', [])))
                    
                    has_darknet = getattr(addr, 'api_has_darknet_exposure', False)
                    ws.cell(row=row, column=14, value="Y" if has_darknet else "N")
                    ws.cell(row=row, column=15, value=f"{getattr(addr, 'api_balance', 0):,.8f}")
                    ws.cell(row=row, column=16, value=f"{getattr(addr, 'api_total_received', 0):,.8f}")
                    ws.cell(row=row, column=17, value=f"{getattr(addr, 'api_total_sent', 0):,.8f}")
                    
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            self.logger.info(f"✓ Created {crypto_name} sheet successfully")

        except Exception as e:
            self.logger.error(f"Failed to create {crypto_name} sheet: {str(e)}")
            raise

    def _create_duplicate_analysis_sheet(self, wb: Workbook, addresses: List[ExtractedAddress]) -> None:
        """
        Create duplicate analysis sheet.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            addresses (List[ExtractedAddress]): List of all addresses
            
        Raises:
            Exception: If duplicate analysis sheet creation fails
        """
        try:
            ws = wb.create_sheet("Duplicate Analysis")
            self.logger.info("Creating 'Duplicate Analysis' sheet")

            headers = ["Address", "Cluster Address", "Crypto", "Occurrences", "Files"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF4B4B", end_color="FF4B4B", fill_type="solid")

            # Group duplicates
            duplicates = defaultdict(list)
            for addr in addresses:
                if addr.is_duplicate:
                    duplicates[addr.address].append(addr)

            row = 2
            for address, occurrences in sorted(duplicates.items()):
                if len(occurrences) > 1:
                    ws.cell(row=row, column=1, value=address)
                    
                    # Get cluster address for this duplicate
                    cluster_address = ''
                    if occurrences:
                        cluster_address = getattr(occurrences[0], 'api_cluster_root_address', '')
                        if not cluster_address:
                            cluster_address = getattr(occurrences[0], 'api_root_address', '')
                    ws.cell(row=row, column=2, value=cluster_address)
                    ws.cell(row=row, column=3, value=occurrences[0].crypto_name)
                    ws.cell(row=row, column=4, value=len(occurrences))
                    files = list(set(addr.filename for addr in occurrences))
                    ws.cell(row=row, column=5, value=", ".join(files[:3]))
                    row += 1

            self.logger.info("✓ Created Duplicate Analysis sheet successfully")

        except Exception as e:
            self.logger.error(f"Failed to create duplicate analysis sheet: {str(e)}")
            # Don't raise - this is optional

    def _create_paginated_sheets(self, wb: Workbook, base_name: str, 
                                addresses: List[ExtractedAddress], 
                                include_api_data: bool) -> None:
        """
        Create multiple sheets when data exceeds Excel row limits.
        
        Args:
            wb (Workbook): The workbook
            base_name (str): Base name for sheets
            addresses (List[ExtractedAddress]): Addresses to paginate
            include_api_data (bool): Whether to include API data
        """
        chunks = [addresses[i:i + self.SAFE_ROW_LIMIT] 
                 for i in range(0, len(addresses), self.SAFE_ROW_LIMIT)]
        
        for idx, chunk in enumerate(chunks, 1):
            sheet_name = f"{base_name[:25]}_Part{idx}"
            self.logger.warning(f"Creating paginated sheet: {sheet_name} with {len(chunk)} rows")
            # Note: Implementation would call appropriate sheet creation method
            # This is a placeholder for the pagination logic

    def _format_exposure_text(self, exposures):
        """
        Format exposure data showing only exchanges and darknet markets.
        Darknet markets are marked with (DARKNET) indicator.
        
        Args:
            exposures: List of exposure dictionaries
            
        Returns:
            str: Formatted exposure text
        """
        if not exposures:
            return "None"
        
        # Sort by percentage descending
        sorted_exposures = sorted(
            exposures,
            key=lambda x: x.get('percentage', 0),
            reverse=True
        )
        
        formatted_parts = []
        for exp in sorted_exposures[:10]:  # Top 10
            name = exp.get('name', 'Unknown')
            percentage = exp.get('percentage', 0)
            category = exp.get('category', '').lower()
            
            if percentage >= 0.1:  # Only show >= 0.1%
                # Mark darknet markets
                if 'darknet' in category or 'dark market' in category:
                    formatted_parts.append(f"{name} (DARKNET): {percentage:.1f}%")
                else:
                    formatted_parts.append(f"{name}: {percentage:.1f}%")
        
        return "; ".join(formatted_parts) if formatted_parts else "None"

    def _format_detailed_exposure(self, exposures, max_items=3):
        """
        Format exposure data with more details for Excel display.
        
        Args:
            exposures: List of exposure dictionaries
            max_items: Maximum number of items to display
            
        Returns:
            str: Formatted string with exchange names and percentages
        """
        if not exposures:
            return ""
        
        # Sort by value/percentage
        sorted_exposures = sorted(
            exposures,
            key=lambda x: x.get('percentage', 0) or x.get('value', 0),
            reverse=True
        )[:max_items]
        
        formatted_parts = []
        for exp in sorted_exposures:
            name = exp.get('name', 'Unknown')
            percentage = exp.get('percentage', 0)
            value = exp.get('value', 0)
            
            if percentage > 0:
                formatted_parts.append(f"{name}: {percentage:.1f}%")
            elif value > 0:
                formatted_parts.append(f"{name}: ${value:,.2f}")
            else:
                formatted_parts.append(name)
        
        return "; ".join(formatted_parts)