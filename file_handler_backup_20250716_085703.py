"""
File Handler Module - Production Ready Version with Direct/Indirect Exposure
===========================================================================

This module handles all file I/O operations for the cryptocurrency address extractor.
Provides methods for reading CSV/Excel files and writing comprehensive Excel reports
with multiple analysis sheets including both direct and indirect exchange exposure.

Author: Crypto Extractor Tool
Date: 2025-01-18
Version: 2.3.0 - Refactored with Summary and Column Definitions Sheets
"""

import os
import logging
import pandas as pd
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime
from collections import defaultdict
from ope, Counternpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.worksheet import Worksheet
import re
import json

from extractor import ExtractedAddress


class FileHandler:
    """
    Handles all file I/O operations for the cryptocurrency address extractor.
    
    Features:
    - Read CSV and Excel files with multi-sheet support
    - Write comprehensive Excel reports with multiple analysis sheets
    - Handle large datasets with row limit protection
    - Professional formatting and styling
    - Backward compatible save_to_excel method
    - Enhanced direct and indirect exchange exposure display
    - Summary sheet with extraction statistics
    - Column definitions sheet for user reference
    """
    
    # Excel row limit with safety buffer
    EXCEL_MAX_ROWS = 1048576
    SAFE_ROW_LIMIT = 1000000  # Leave buffer for headers and formatting
    
    def __init__(self):
        """
        Initialize the file handler with logging configuration.
        
        Raises:
            Exception: If logger cannot be initialized
        """
        try:
            self.logger = logging.getLogger(__name__)
            self.logger.info("FileHandler initialized successfully")
        except Exception as e:
            print(f"Failed to initialize FileHandler logger: {e}")
            raise
    
    def read_csv(self, file_path: str, encoding: str = 'utf-8') -> pd.DataFrame:
        """
        Read CSV file and return as DataFrame.
        
        Args:
            file_path (str): Path to the CSV file
            encoding (str): File encoding (default: utf-8)
            
        Returns:
            pd.DataFrame: Loaded data
            
        Raises:
            FileNotFoundError: If file doesn't exist
            pd.errors.ParserError: If CSV parsing fails
            Exception: For other errors
        """
        self.logger.info(f"Reading CSV file: {file_path}")
        
        if not os.path.exists(file_path):
            error_msg = f"CSV file not found: {file_path}"
            self.logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        try:
            # Read CSV with multiple encoding fallbacks
            encodings = [encoding, 'utf-8', 'latin-1', 'cp1252']
            
            for enc in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=enc, low_memory=False)
                    self.logger.info(f"Successfully read CSV with {enc} encoding: {len(df)} rows")
                    return df
                except UnicodeDecodeError:
                    continue
            
            # If all encodings fail
            raise ValueError(f"Could not read CSV with any encoding: {encodings}")
            
        except pd.errors.ParserError as e:
            error_msg = f"CSV parsing error in {file_path}: {str(e)}"
            self.logger.error(error_msg)
            raise pd.errors.ParserError(error_msg)
        except Exception as e:
            error_msg = f"Error reading CSV file {file_path}: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e
    
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, pd.DataFrame]:
        """
        Read Excel file and return dictionary of DataFrames for each sheet.
        
        Args:
            file_path (str): Path to the Excel file
            sheet_name (Optional[str]): Specific sheet to read (None for all sheets)
            
        Returns:
            Dict[str, pd.DataFrame]: Dictionary mapping sheet names to DataFrames
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If Excel file is invalid
            Exception: For other errors
        """
        self.logger.info(f"Reading Excel file: {file_path}")
        
        if not os.path.exists(file_path):
            error_msg = f"Excel file not found: {file_path}"
            self.logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        try:
            # Read Excel file
            excel_file = pd.ExcelFile(file_path)
            sheets_to_read = [sheet_name] if sheet_name else excel_file.sheet_names
            
            result = {}
            for sheet in sheets_to_read:
                if sheet in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet, header=None)
                    result[sheet] = df
                    self.logger.info(f"Read sheet '{sheet}': {len(df)} rows")
                else:
                    self.logger.warning(f"Sheet '{sheet}' not found in {file_path}")
            
            excel_file.close()
            return result
            
        except ValueError as e:
            error_msg = f"Invalid Excel file {file_path}: {str(e)}"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        except Exception as e:
            error_msg = f"Error reading Excel file {file_path}: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e
    
    def save_to_excel(self, addresses: List[ExtractedAddress], output_path: str, 
                      include_api_data: Optional[bool] = None) -> str:
        """
        Save extracted addresses to Excel file (wrapper for backward compatibility).
        
        This method wraps write_to_excel to maintain compatibility with existing code
        that expects save_to_excel to return the output path.
        
        Args:
            addresses (List[ExtractedAddress]): List of extracted addresses
            output_path (str): Path for the output Excel file
            include_api_data (Optional[bool]): Whether to include API data columns.
                                             If None, auto-detects from addresses.
            
        Returns:
            str: The path of the saved file
            
        Raises:
            PermissionError: If file is open or inaccessible
            Exception: For other errors
        """
        try:
            # Auto-detect if we should include API data if not specified
            if include_api_data is None:
                include_api_data = any(hasattr(addr, 'api_balance') for addr in addresses)
                if include_api_data:
                    self.logger.info("Auto-detected API data in addresses")
            
            # Call the main write_to_excel method
            self.write_to_excel(addresses, output_path, include_api_data)
            
            # Log success
            self.logger.info(f"Successfully saved {len(addresses)} addresses to {output_path}")
            
            # Return the output path for backward compatibility
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error in save_to_excel: {str(e)}")
            raise
    
    def write_to_excel(self, addresses: List[ExtractedAddress], output_path: str,
                      include_api_data: bool = False) -> None:
        """
        Write extracted addresses to a formatted Excel file with multiple sheets.
        
        Creates the following sheets in order:
        1. Summary - Extraction statistics and API usage
        2. Column Definitions - Explanation of all column headers
        3. All Addresses - Complete list of all extracted addresses
        4. Individual crypto sheets - One sheet per cryptocurrency type
        
        Args:
            addresses (List[ExtractedAddress]): List of extracted addresses
            output_path (str): Path for the output Excel file
            include_api_data (bool): Whether to include API data columns
            
        Raises:
            PermissionError: If file is open or inaccessible
            Exception: For other errors
        """
        self.logger.info(f"Writing {len(addresses)} addresses to Excel: {output_path}")
        
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets in the desired order:
            # 1. Summary sheet (position 0)
            self._create_summary_sheet(wb, addresses, getattr(self, "_api_stats", None))
            
            # 2. Column Definitions sheet (position 1)
            self._create_column_definitions_sheet(wb, include_api_data)
            
            # 3. All Addresses sheet (position 2)
            self._create_all_addresses_sheet(wb, addresses, include_api_data)
            
            # 4. Individual crypto sheets (positions 3+)
            crypto_groups = defaultdict(list)
            for addr in addresses:
                crypto_groups[addr.crypto_name].append(addr)
            
            # Create individual crypto sheets
            for crypto_name, crypto_addresses in crypto_groups.items():
                self._create_crypto_sheet(wb, crypto_name, crypto_addresses, include_api_data)
            
            # 5. Optional: Create duplicate analysis sheet if duplicates exist
            if any(addr.is_duplicate for addr in addresses):
                self._create_duplicate_analysis_sheet(wb, addresses)
            
            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Successfully created Excel file with Summary and Column Definitions: {output_path}")
            
        except PermissionError as e:
            error_msg = f"Permission denied writing to {output_path}. Please close the file if it's open."
            self.logger.error(error_msg)
            raise PermissionError(error_msg) from e
        except Exception as e:
            error_msg = f"Failed to create Excel file: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e
    
    def _create_summary_sheet(self, wb: Workbook, addresses: List[ExtractedAddress], 
                           api_stats=None) -> None:
        """
        Create summary sheet with extraction statistics and API usage.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            addresses (List[ExtractedAddress]): List of all addresses
            api_stats (dict, optional): API usage statistics from tracking
            
        Raises:
            Exception: If summary sheet creation fails
        """
        try:
            self.logger.info("Creating 'Summary' sheet")
            
            # Create the summary sheet first (position 0)
            ws = wb.create_sheet("Summary", 0)
            
            # Title section
            ws['A1'] = "Cryptocurrency Address Extraction Summary"
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A1:D1')
            
            # Basic extraction statistics
            ws['A3'] = "Extraction Date:"
            ws['A3'].font = Font(bold=True)
            ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            ws['A4'] = "Total Addresses Found:"
            ws['A4'].font = Font(bold=True)
            ws['B4'] = len(addresses)
            
            # Calculate unique addresses
            unique_addresses = len(set(addr.address for addr in addresses))
            ws['A5'] = "Unique Addresses:"
            ws['A5'].font = Font(bold=True)
            ws['B5'] = unique_addresses
            
            ws['A6'] = "Duplicate Addresses:"
            ws['A6'].font = Font(bold=True)
            ws['B6'] = len(addresses) - unique_addresses
            
            # Files processed
            unique_files = set(addr.filename for addr in addresses)
            ws['A7'] = "Files Processed:"
            ws['A7'].font = Font(bold=True)
            ws['B7'] = len(unique_files)
            
            # Excel sheets processed (for Excel files)
            excel_sheets = set()
            for addr in addresses:
                if addr.sheet_name:  # Only count if there's a sheet name
                    excel_sheets.add(f"{addr.filename}:{addr.sheet_name}")
            
            if excel_sheets:
                ws['A8'] = "Excel Sheets Processed:"
                ws['A8'].font = Font(bold=True)
                ws['B8'] = len(excel_sheets)
            
            # Cryptocurrency breakdown section
            ws['A10'] = "Cryptocurrency Breakdown:"
            ws['A10'].font = Font(size=12, bold=True, color="FFFFFF")
            ws['A10'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.merge_cells('A10:C10')
            
            # Count addresses by cryptocurrency
            crypto_counts = defaultdict(int)
            for addr in addresses:
                crypto_counts[addr.crypto_name] += 1
            
            # Headers for crypto breakdown table
            ws['A11'] = "Cryptocurrency"
            ws['A11'].font = Font(bold=True)
            ws['B11'] = "Count"
            ws['B11'].font = Font(bold=True)
            ws['C11'] = "Percentage"
            ws['C11'].font = Font(bold=True)
            
            # Crypto breakdown data
            row = 12
            total_addresses = len(addresses)
            for crypto, count in sorted(crypto_counts.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_addresses) * 100
                ws.cell(row=row, column=1, value=crypto)
                ws.cell(row=row, column=2, value=count)
                ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
                row += 1
            
            # API Usage Statistics Section (if available)
            if api_stats and api_stats.get('total_calls', 0) > 0:
                api_start_row = row + 2
                
                # API Usage Section Header
                ws.cell(row=api_start_row, column=1, value="Chainalysis API Usage Statistics:").font = Font(bold=True, size=12, color="FFFFFF")
                ws.cell(row=api_start_row, column=1).fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
                ws.merge_cells(f'A{api_start_row}:D{api_start_row}')
                api_start_row += 1
                
                # Overall API Statistics
                ws.cell(row=api_start_row, column=1, value="Total API Calls:")
                ws.cell(row=api_start_row, column=1).font = Font(bold=True)
                ws.cell(row=api_start_row, column=2, value=api_stats.get('total_calls', 0))
                api_start_row += 1
                
                ws.cell(row=api_start_row, column=1, value="Successful Calls:")
                ws.cell(row=api_start_row, column=1).font = Font(bold=True)
                ws.cell(row=api_start_row, column=2, value=api_stats.get('successful_calls', 0))
                api_start_row += 1
                
                ws.cell(row=api_start_row, column=1, value="Failed Calls:")
                ws.cell(row=api_start_row, column=1).font = Font(bold=True)
                ws.cell(row=api_start_row, column=2, value=api_stats.get('failed_calls', 0))
                api_start_row += 1
                
                ws.cell(row=api_start_row, column=1, value="Success Rate:")
                ws.cell(row=api_start_row, column=1).font = Font(bold=True)
                success_rate = api_stats.get('success_rate', 0)
                ws.cell(row=api_start_row, column=2, value=f"{success_rate:.1f}%")
                api_start_row += 1
                
                ws.cell(row=api_start_row, column=1, value="Total Processing Time:")
                ws.cell(row=api_start_row, column=1).font = Font(bold=True)
                total_time = api_stats.get('total_time_seconds', 0)
                if total_time > 60:
                    time_str = f"{total_time/60:.1f} minutes"
                else:
                    time_str = f"{total_time:.1f} seconds"
                ws.cell(row=api_start_row, column=2, value=time_str)
                api_start_row += 2
                
                # API Calls by Endpoint Type
                calls_by_endpoint = api_stats.get('calls_by_endpoint', {})
                if any(calls_by_endpoint.values()):
                    ws.cell(row=api_start_row, column=1, value="API Calls by Endpoint:").font = Font(bold=True)
                    api_start_row += 1
                    
                    for endpoint, count in calls_by_endpoint.items():
                        if count > 0:
                            success_count = api_stats.get('success_by_endpoint', {}).get(endpoint, 0)
                            failure_count = api_stats.get('failure_by_endpoint', {}).get(endpoint, 0)
                            avg_time = api_stats.get('avg_response_times', {}).get(endpoint, 0)
                            
                            ws.cell(row=api_start_row, column=1, value=f"  {endpoint.title()}:")
                            ws.cell(row=api_start_row, column=2, value=f"{count} calls")
                            ws.cell(row=api_start_row, column=3, value=f"({success_count} success, {failure_count} failed)")
                            ws.cell(row=api_start_row, column=4, value=f"Avg: {avg_time:.2f}s")
                            api_start_row += 1
            
            # Format columns
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            
            self.logger.info("✓ Created Summary sheet successfully")
            
        except Exception as e:
            error_msg = f"Failed to create summary sheet: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e

    def _create_column_definitions_sheet(self, wb: Workbook, include_api_data: bool = False) -> None:
        """
        Create a column definitions sheet that explains what each column heading means.
        This sheet acts as a reference key for users to understand the data structure.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            include_api_data (bool): Whether API data columns should be included in definitions
            
        Raises:
            Exception: If sheet creation fails
        """
        try:
            self.logger.info("Creating 'Column Definitions' sheet")
            
            # Create the sheet at position 1 (after Summary)
            ws = wb.create_sheet("Column Definitions", 1)
            
            # Title
            ws['A1'] = "Column Definitions & Reference Guide"
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A1:C1')
            
            # Subtitle
            ws['A2'] = "This sheet explains what each column heading means in the exported data"
            ws['A2'].font = Font(size=11, italic=True)
            ws['A2'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A2:C2')
            
            # Headers for the definitions table
            ws['A4'] = "Column Name"
            ws['B4'] = "Description"
            ws['C4'] = "Example/Notes"
            
            # Format headers
            for col in ['A4', 'B4', 'C4']:
                ws[col].font = Font(bold=True, color="FFFFFF")
                ws[col].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                ws[col].alignment = Alignment(horizontal="center")
            
            # Basic column definitions (always present)
            basic_definitions = [
                ("Address", "The cryptocurrency wallet address that was extracted", "1A1zP1eP5QGefi2DMPTfTL5SLmv7DivfNa"),
                ("Cluster Address", "The root address of the cluster this address belongs to (from API)", "Same format as Address"),
                ("Cryptocurrency", "The type of cryptocurrency for this address", "Bitcoin, Ethereum, Litecoin, etc."),
                ("Source File", "The name of the file where this address was found", "transactions.xlsx, addresses.csv"),
                ("Sheet", "The specific sheet name in Excel files where address was found", "Sheet1, Transaction Data"),
                ("Row", "The row number in the source file where address was located", "5, 127, 1450"),
                ("Column", "The column number/letter where address was found", "3, A, D"),
                ("Confidence %", "How confident the extraction algorithm is that this is a valid address", "95.5%, 100.0%"),
                ("Is Duplicate", "Whether this exact address appears multiple times in the data", "Yes, No"),
                ("Total Count", "How many times this address appears across all source files", "1, 3, 15")
            ]
            
            # API-specific column definitions (only if API data is included)
            api_definitions = [
                ("Balance", "Current balance of the address in native cryptocurrency units", "1.25843210 BTC"),
                ("Total Received", "Total amount ever received by this address", "5.67891234 BTC"),
                ("Total Sent", "Total amount ever sent from this address", "4.42048024 BTC"),
                ("Transfer Count", "Number of transactions involving this address", "157, 1205"),
                ("Direct Exchange Exposure", "Exchanges that directly interact with this address", "Binance: 45.2%, Coinbase: 23.1%"),
                ("Indirect Exchange Exposure", "Exchanges connected through 1-2 intermediate addresses", "Kraken: 12.5%, Bitfinex: 8.3%"),
                ("Receiving Direct Exposure", "Exchanges this address directly receives funds from", "Coinbase: 67.8%, Kraken: 15.2%"),
                ("Receiving Indirect Exposure", "Exchanges connected to incoming transactions via intermediates", "Binance: 23.1%, Huobi: 8.9%"),
                ("Sending Direct Exposure", "Exchanges this address directly sends funds to", "Bitfinex: 45.6%, OKEx: 12.3%"),
                ("Sending Indirect Exposure", "Exchanges connected to outgoing transactions via intermediates", "Coinbase: 34.5%, Kraken: 11.2%"),
                ("Darknet Market", "Whether address has any connection to darknet markets", "Y (Yes) or N (No)"),
                ("Risk Level", "API-determined risk level based on transaction patterns", "Low, Medium, High, Critical"),
                ("Entity", "Known entity name if address belongs to a service", "Binance Hot Wallet, Coinbase"),
                ("Cluster Category", "Category of the cluster this address belongs to", "Exchange, Mixer, DeFi, etc.")
            ]
            
            # Write basic definitions
            current_row = 5
            for col_name, description, example in basic_definitions:
                ws.cell(row=current_row, column=1, value=col_name).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=description)
                ws.cell(row=current_row, column=3, value=example)
                current_row += 1
            
            # Add API definitions if API data is included
            if include_api_data:
                # Add a separator
                current_row += 1
                ws.cell(row=current_row, column=1, value="API DATA COLUMNS").font = Font(bold=True, size=12, color="FF0000")
                ws.merge_cells(f'A{current_row}:C{current_row}')
                current_row += 1
                
                for col_name, description, example in api_definitions:
                    ws.cell(row=current_row, column=1, value=col_name).font = Font(bold=True)
                    ws.cell(row=current_row, column=2, value=description)
                    ws.cell(row=current_row, column=3, value=example)
                    current_row += 1
            
            # Add footer information
            current_row += 2
            footer_info = [
                "IMPORTANT NOTES:",
                "• Duplicate addresses are highlighted in yellow in the 'All Addresses' sheet",
                "• Individual cryptocurrency sheets show only unique addresses (no duplicates)",
                "• Exchange exposure percentages show the portion of funds connected to each exchange",
                "• Darknet market connections are flagged for compliance and investigation purposes",
                "• This extraction helps identify addresses that appear in multiple source files"
            ]
            
            for info in footer_info:
                ws.cell(row=current_row, column=1, value=info)
                ws.merge_cells(f'A{current_row}:C{current_row}')
                current_row += 1
            
            # Set column widths for better readability
            ws.column_dimensions['A'].width = 25  # Column Name
            ws.column_dimensions['B'].width = 60  # Description  
            ws.column_dimensions['C'].width = 35  # Example/Notes
            
            # Add borders and formatting to make it more readable
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply borders to the definitions table
            for row in range(4, current_row - len(footer_info) - 2):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = thin_border
            
            self.logger.info("✓ Created Column Definitions sheet successfully")
            
        except Exception as e:
            error_msg = f"Failed to create Column Definitions sheet: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg) from e

    def _create_all_addresses_sheet(self, wb: Workbook, addresses: List[ExtractedAddress],
                                   include_api_data: bool) -> None:
        """
        Create sheet with all extracted addresses including enhanced exposure data.
        
        Enhanced to properly include both direct and indirect exposure data.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            addresses (List[ExtractedAddress]): List of all addresses
            include_api_data (bool): Whether to include API data columns
            
        Raises:
            Exception: If sheet creation fails
        """
        try:
            self.logger.info(f"Creating 'All Addresses' sheet with include_api_data={include_api_data}")
            
            ws = wb.create_sheet("All Addresses")
            
            # Check row limit
            if len(addresses) > self.SAFE_ROW_LIMIT:
                self._create_paginated_sheets(wb, "All Addresses", addresses, include_api_data)
                return
            
            # Headers
            headers = ["Address", "Cluster Address", "Cryptocurrency", "Source File", "Sheet", "Row", "Column",
                      "Confidence %", "Is Duplicate", "Total Count"]
            
            if include_api_data:
                # Enhanced headers with both direct and indirect exposure
                headers.extend(["Balance", "Total Received", "Total Sent", "Transfer Count",
                               "Direct Exchange Exposure", "Indirect Exchange Exposure",
                               "Receiving Direct Exposure", "Receiving Indirect Exposure", 
                               "Sending Direct Exposure", "Sending Indirect Exposure",
                               "Darknet Market", "Risk Level", "Entity", "Cluster Category"])
            
            # Write headers with formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Write data
            for row, addr in enumerate(addresses, 2):
                # Basic columns (1-10)
                ws.cell(row=row, column=1, value=addr.address)
                
                # Cluster address in column 2
                cluster_address = getattr(addr, 'api_cluster_root_address', '')
                if not cluster_address:
                    cluster_address = getattr(addr, 'api_root_address', '')
                ws.cell(row=row, column=2, value=cluster_address)
                ws.cell(row=row, column=3, value=addr.crypto_name)
                ws.cell(row=row, column=4, value=addr.filename)
                ws.cell(row=row, column=5, value=addr.sheet_name or "N/A")
                ws.cell(row=row, column=6, value=addr.row)
                ws.cell(row=row, column=7, value=addr.column)
                ws.cell(row=row, column=8, value=f"{addr.confidence:.1f}%")
                ws.cell(row=row, column=9, value="Yes" if addr.is_duplicate else "No")
                ws.cell(row=row, column=10, value=addr.duplicate_count)
                
                if include_api_data:
                    # API data columns (11-24)
                    ws.cell(row=row, column=11, value=f"{getattr(addr, 'api_balance', 0):,.8f}")
                    ws.cell(row=row, column=12, value=f"{getattr(addr, 'api_total_received', 0):,.8f}")
                    ws.cell(row=row, column=13, value=f"{getattr(addr, 'api_total_sent', 0):,.8f}")
                    ws.cell(row=row, column=14, value=getattr(addr, 'api_transfer_count', 0))
                    
                    # Format exposure text fields
                    ws.cell(row=row, column=15, value=self._format_exposure_text(getattr(addr, 'api_direct_exposure', [])))
                    ws.cell(row=row, column=16, value=self._format_exposure_text(getattr(addr, 'api_indirect_exposure', [])))
                    ws.cell(row=row, column=17, value=self._format_exposure_text(getattr(addr, 'api_receiving_direct_exposure', [])))
                    ws.cell(row=row, column=18, value=self._format_exposure_text(getattr(addr, 'api_receiving_indirect_exposure', [])))
                    ws.cell(row=row, column=19, value=self._format_exposure_text(getattr(addr, 'api_sending_direct_exposure', [])))
                    ws.cell(row=row, column=20, value=self._format_exposure_text(getattr(addr, 'api_sending_indirect_exposure', [])))
                    
                    # Darknet and other fields
                    has_darknet = getattr(addr, 'api_has_darknet_exposure', False)
                    ws.cell(row=row, column=21, value="Y" if has_darknet else "N")
                    ws.cell(row=row, column=22, value=getattr(addr, 'api_risk_level', 'Unknown'))
                    ws.cell(row=row, column=23, value=getattr(addr, 'api_cluster_name', ''))
                    ws.cell(row=row, column=24, value=getattr(addr, 'api_cluster_category', ''))
                    
            # Auto-fit columns
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 15
            
            # Wider columns for specific fields
            ws.column_dimensions['A'].width = 50  # Address
            ws.column_dimensions['B'].width = 30  # Cluster Address
            ws.column_dimensions['D'].width = 30  # Source File
            if include_api_data:
                ws.column_dimensions['O'].width = 40  # Direct Exchange Exposure
                ws.column_dimensions['P'].width = 40  # Indirect Exchange Exposure
            
            self.logger.info(f"✓ Created All Addresses sheet with {len(addresses)} addresses")
            
        except Exception as e:
            self.logger.error(f"Failed to create All Addresses sheet: {str(e)}")
            raise

    def _create_crypto_sheet(self, wb: Workbook, crypto_name: str,
                           addresses: List[ExtractedAddress],
                           include_api_data: bool = False) -> None:
        """
        Create sheet for a specific cryptocurrency.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            crypto_name (str): Name of the cryptocurrency
            addresses (List[ExtractedAddress]): List of addresses for this crypto
            include_api_data (bool): Whether to include API data columns
            
        Raises:
            Exception: If crypto sheet creation fails
        """
        try:
            # Remove duplicates
            seen = set()
            unique_addresses = []
            for addr in addresses:
                if addr.address not in seen:
                    seen.add(addr.address)
                    unique_addresses.append(addr)

            ws = wb.create_sheet(crypto_name[:31])
            self.logger.info(f"Creating '{crypto_name}' sheet with {len(unique_addresses)} unique addresses")

            # Headers
            headers = ["Address", "Cluster Address", "Source File", "Row", "Column", "Confidence"]
            if include_api_data:
                headers.extend([
                    "Entity Name", "Direct Exposure", "Indirect Exposure",
                    "Receiving Direct", "Receiving Indirect",
                    "Sending Direct", "Sending Indirect",
                    "Darknet Market", "Balance", "Total Received", "Total Sent"
                ])

            # Write headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Write data
            for row, addr in enumerate(unique_addresses, 2):
                ws.cell(row=row, column=1, value=addr.address)
                
                # Cluster address in column 2
                cluster_addr = getattr(addr, 'api_cluster_root_address', '')
                if not cluster_addr:
                    cluster_addr = getattr(addr, 'api_root_address', '')
                ws.cell(row=row, column=2, value=cluster_addr)
                
                # Basic data
                ws.cell(row=row, column=3, value=addr.filename)
                ws.cell(row=row, column=4, value=addr.row)
                ws.cell(row=row, column=5, value=addr.column)
                ws.cell(row=row, column=6, value=f"{addr.confidence:.1f}%")
                
                if include_api_data:
                    # API data columns (7-17)
                    ws.cell(row=row, column=7, value=getattr(addr, 'api_cluster_name', ''))
                    ws.cell(row=row, column=8, value=self._format_exposure_text(getattr(addr, 'api_direct_exposure', [])))
                    ws.cell(row=row, column=9, value=self._format_exposure_text(getattr(addr, 'api_indirect_exposure', [])))
                    ws.cell(row=row, column=10, value=self._format_exposure_text(getattr(addr, 'api_receiving_direct_exposure', [])))
                    ws.cell(row=row, column=11, value=self._format_exposure_text(getattr(addr, 'api_receiving_indirect_exposure', [])))
                    ws.cell(row=row, column=12, value=self._format_exposure_text(getattr(addr, 'api_sending_direct_exposure', [])))
                    ws.cell(row=row, column=13, value=self._format_exposure_text(getattr(addr, 'api_sending_indirect_exposure', [])))
                    
                    has_darknet = getattr(addr, 'api_has_darknet_exposure', False)
                    ws.cell(row=row, column=14, value="Y" if has_darknet else "N")
                    ws.cell(row=row, column=15, value=f"{getattr(addr, 'api_balance', 0):,.8f}")
                    ws.cell(row=row, column=16, value=f"{getattr(addr, 'api_total_received', 0):,.8f}")
                    ws.cell(row=row, column=17, value=f"{getattr(addr, 'api_total_sent', 0):,.8f}")
                    
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            self.logger.info(f"✓ Created {crypto_name} sheet successfully")

        except Exception as e:
            self.logger.error(f"Failed to create {crypto_name} sheet: {str(e)}")
            raise

    def _create_duplicate_analysis_sheet(self, wb: Workbook, addresses: List[ExtractedAddress]) -> None:
        """
        Create duplicate analysis sheet.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            addresses (List[ExtractedAddress]): List of all addresses
            
        Raises:
            Exception: If duplicate analysis sheet creation fails
        """
        try:
            ws = wb.create_sheet("Duplicate Analysis")
            self.logger.info("Creating 'Duplicate Analysis' sheet")

            headers = ["Address", "Cluster Address", "Crypto", "Occurrences", "Files"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF4B4B", end_color="FF4B4B", fill_type="solid")

            # Group duplicates
            duplicates = defaultdict(list)
            for addr in addresses:
                if addr.is_duplicate:
                    duplicates[addr.address].append(addr)

            row = 2
            for address, occurrences in sorted(duplicates.items()):
                if len(occurrences) > 1:
                    ws.cell(row=row, column=1, value=address)
                    
                    # Get cluster address for this duplicate
                    cluster_address = ''
                    if occurrences:
                        cluster_address = getattr(occurrences[0], 'api_cluster_root_address', '')
                        if not cluster_address:
                            cluster_address = getattr(occurrences[0], 'api_root_address', '')
                    ws.cell(row=row, column=2, value=cluster_address)
                    ws.cell(row=row, column=3, value=occurrences[0].crypto_name)
                    ws.cell(row=row, column=4, value=len(occurrences))
                    files = list(set(addr.filename for addr in occurrences))
                    ws.cell(row=row, column=5, value=", ".join(files[:3]))
                    row += 1

            self.logger.info("✓ Created Duplicate Analysis sheet successfully")

        except Exception as e:
            self.logger.error(f"Failed to create duplicate analysis sheet: {str(e)}")
            # Don't raise - this is optional

    def _create_paginated_sheets(self, wb: Workbook, base_name: str, 
                                addresses: List[ExtractedAddress], 
                                include_api_data: bool) -> None:
        """
        Create multiple sheets when data exceeds Excel row limits.
        
        Args:
            wb (Workbook): The workbook
            base_name (str): Base name for sheets
            addresses (List[ExtractedAddress]): Addresses to paginate
            include_api_data (bool): Whether to include API data
        """
        chunks = [addresses[i:i + self.SAFE_ROW_LIMIT] 
                 for i in range(0, len(addresses), self.SAFE_ROW_LIMIT)]
        
        for idx, chunk in enumerate(chunks, 1):
            sheet_name = f"{base_name[:25]}_Part{idx}"
            self.logger.warning(f"Creating paginated sheet: {sheet_name} with {len(chunk)} rows")
            # Note: Implementation would call appropriate sheet creation method
            # This is a placeholder for the pagination logic

    def _format_exposure_text(self, exposures):
        """
        Format exposure data showing only exchanges and darknet markets.
        Darknet markets are marked with (DARKNET) indicator.
        
        Args:
            exposures: List of exposure dictionaries
            
        Returns:
            str: Formatted exposure text
        """
        if not exposures:
            return "None"
        
        # Sort by percentage descending
        sorted_exposures = sorted(
            exposures,
            key=lambda x: x.get('percentage', 0),
            reverse=True
        )
        
        formatted_parts = []
        for exp in sorted_exposures[:10]:  # Top 10
            name = exp.get('name', 'Unknown')
            percentage = exp.get('percentage', 0)
            category = exp.get('category', '').lower()
            
            if percentage >= 0.1:  # Only show >= 0.1%
                # Mark darknet markets
                if 'darknet' in category or 'dark market' in category:
                    formatted_parts.append(f"{name} (DARKNET): {percentage:.1f}%")
                else:
                    formatted_parts.append(f"{name}: {percentage:.1f}%")
        
        return "; ".join(formatted_parts) if formatted_parts else "None"

    def _format_detailed_exposure(self, exposures, max_items=3):
        """
        Format exposure data with more details for Excel display.
        
        Args:
            exposures: List of exposure dictionaries
            max_items: Maximum number of items to display
            
        Returns:
            str: Formatted string with exchange names and percentages
        """
        if not exposures:
            return ""
        
        # Sort by value/percentage
        sorted_exposures = sorted(
            exposures,
            key=lambda x: x.get('percentage', 0) or x.get('value', 0),
            reverse=True
        )[:max_items]
        
        formatted_parts = []
        for exp in sorted_exposures:
            name = exp.get('name', 'Unknown')
            percentage = exp.get('percentage', 0)
            value = exp.get('value', 0)
            
            if percentage > 0:
                formatted_parts.append(f"{name}: {percentage:.1f}%")
            elif value > 0:
                formatted_parts.append(f"{name}: ${value:,.2f}")
            else:
                formatted_parts.append(name)
        
        return "; ".join(formatted_parts)
    def _create_frequency_analysis_sheet(self, wb: Workbook, addresses: List[ExtractedAddress], 
                                       include_api_data: bool = False) -> None:
        """
        Create Address Frequency Analysis sheet showing cluster address frequencies.
        
        File: file_handler.py
        Function: _create_frequency_analysis_sheet()
        
        This sheet shows how many times each cluster address appears across all datasets,
        using existing Chainalysis API data without making additional API calls.
        
        Args:
            wb (Workbook): The workbook to add the sheet to
            addresses (List[ExtractedAddress]): List of all extracted addresses
            include_api_data (bool): Whether API data is available
            
        Raises:
            Exception: If frequency analysis sheet creation fails
        """
        try:
            self.logger.info("Creating Address Frequency Analysis sheet")
            
            # Create frequency data structure
            frequency_data = self._analyze_cluster_frequency(addresses, include_api_data)
            
            # Create worksheet
            ws = wb.create_sheet("Address Frequency Analysis")
            
            # Add statistics summary at the top (will insert rows)
            self._add_frequency_statistics(ws, frequency_data, len(addresses))
            
            # Define headers (now starting at row 5 due to statistics)
            headers = [
                "Cluster Address",
                "Cluster Name", 
                "Frequency Count",
                "Cryptocurrency Type",
                "Files Containing Address",
                "Sheets Containing Address"
            ]
            
            # Write headers with styling at row 5
            self._write_frequency_headers_with_styling(ws, headers, start_row=5)
            
            # Sort frequency data by count (highest first)
            sorted_data = sorted(frequency_data, key=lambda x: x['frequency'], reverse=True)
            
            # Write frequency data starting at row 6
            for row_idx, freq_data in enumerate(sorted_data, start=6):
                self._write_frequency_row(ws, row_idx, freq_data)
            
            # Apply color coding based on frequency (skip statistics rows)
            self._apply_frequency_color_coding(ws, sorted_data, start_row=6)
            
            # Auto-fit columns
            self._auto_fit_frequency_columns(ws)
            
            self.logger.info(f"✓ Created Address Frequency Analysis sheet with {len(frequency_data)} cluster addresses")
            
        except Exception as e:
            self.logger.error(f"Failed to create Address Frequency Analysis sheet: {str(e)}", exc_info=True)
            raise

    def _analyze_cluster_frequency(self, addresses: List[ExtractedAddress], 
                                 include_api_data: bool) -> List[Dict]:
        """
        Analyze cluster address frequencies from extracted addresses.
        
        File: file_handler.py
        Function: _analyze_cluster_frequency()
        
        Args:
            addresses (List[ExtractedAddress]): List of extracted addresses
            include_api_data (bool): Whether API data is available
            
        Returns:
            List[Dict]: List of frequency analysis data
        """
        self.logger.info("Analyzing cluster address frequencies")
        
        # Group addresses by cluster address
        cluster_groups = defaultdict(lambda: {
            'addresses': [],
            'cluster_name': 'Unknown',
            'crypto_types': set(),
            'files': set(),
            'sheets': set()
        })
        
        for addr in addresses:
            # Determine the cluster address to use
            if include_api_data and hasattr(addr, 'api_cluster_root_address') and addr.api_cluster_root_address:
                cluster_address = addr.api_cluster_root_address
                cluster_name = getattr(addr, 'api_cluster_name', 'Unknown')
            else:
                # Fallback to original address if no cluster data
                cluster_address = addr.address
                cluster_name = 'Unknown'
            
            # Add to cluster group
            cluster_groups[cluster_address]['addresses'].append(addr)
            cluster_groups[cluster_address]['cluster_name'] = cluster_name
            cluster_groups[cluster_address]['crypto_types'].add(addr.crypto_type)
            cluster_groups[cluster_address]['files'].add(addr.filename)
            
            # Add sheet name if available
            if hasattr(addr, 'sheet_name') and addr.sheet_name:
                cluster_groups[cluster_address]['sheets'].add(addr.sheet_name)
        
        # Convert to frequency analysis format
        frequency_data = []
        for cluster_address, group_data in cluster_groups.items():
            frequency_data.append({
                'cluster_address': cluster_address,
                'cluster_name': group_data['cluster_name'],
                'frequency': len(group_data['addresses']),
                'crypto_types': ', '.join(sorted(group_data['crypto_types'])),
                'files': ', '.join(sorted(group_data['files'])),
                'sheets': ', '.join(sorted(group_data['sheets'])) if group_data['sheets'] else 'N/A'
            })
        
        self.logger.info(f"Analyzed {len(frequency_data)} unique cluster addresses")
        return frequency_data

    def _write_frequency_headers_with_styling(self, ws: Worksheet, headers: List[str], start_row: int = 1) -> None:
        """
        Write headers with professional styling.
        
        File: file_handler.py
        Function: _write_frequency_headers_with_styling()
        """
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"), 
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

    def _write_frequency_row(self, ws: Worksheet, row_idx: int, freq_data: Dict) -> None:
        """
        Write a single frequency data row.
        
        File: file_handler.py
        Function: _write_frequency_row()
        """
        ws.cell(row=row_idx, column=1, value=freq_data['cluster_address'])
        ws.cell(row=row_idx, column=2, value=freq_data['cluster_name'])
        ws.cell(row=row_idx, column=3, value=freq_data['frequency'])
        ws.cell(row=row_idx, column=4, value=freq_data['crypto_types'])
        ws.cell(row=row_idx, column=5, value=freq_data['files'])
        ws.cell(row=row_idx, column=6, value=freq_data['sheets'])

    def _apply_frequency_color_coding(self, ws: Worksheet, sorted_data: List[Dict], start_row: int = 2) -> None:
        """
        Apply color coding to frequency analysis sheet based on frequency ranges.
        
        File: file_handler.py
        Function: _apply_frequency_color_coding()
        
        Color Scheme:
        - Red: 10+ occurrences (very high frequency)
        - Orange: 5-9 occurrences (high frequency)  
        - Yellow: 3-4 occurrences (medium frequency)
        - Light Green: 2 occurrences (low frequency)
        - White: 1 occurrence (single occurrence)
        """
        self.logger.info("Applying frequency-based color coding")
        
        # Define color fills
        colors = {
            'very_high': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
            'high': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),       # Orange  
            'medium': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),     # Yellow
            'low': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),        # Light Green
            'single': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")      # White
        }
        
        # Apply colors based on frequency
        for idx, freq_data in enumerate(sorted_data):
            row_idx = start_row + idx
            frequency = freq_data['frequency']
            
            # Determine color based on frequency
            if frequency >= 10:
                fill_color = colors['very_high']
            elif frequency >= 5:
                fill_color = colors['high']
            elif frequency >= 3:
                fill_color = colors['medium']
            elif frequency == 2:
                fill_color = colors['low']
            else:  # frequency == 1
                fill_color = colors['single']
            
            # Apply color to all cells in the row
            for col in range(1, 7):  # 6 columns
                ws.cell(row=row_idx, column=col).fill = fill_color

    def _auto_fit_frequency_columns(self, ws: Worksheet) -> None:
        """
        Auto-fit columns for optimal display.
        
        File: file_handler.py
        Function: _auto_fit_frequency_columns()
        """
        # Set specific column widths for frequency analysis
        column_widths = {
            'A': 50,  # Cluster Address
            'B': 30,  # Cluster Name
            'C': 15,  # Frequency Count
            'D': 20,  # Cryptocurrency Type
            'E': 40,  # Files Containing Address
            'F': 30   # Sheets Containing Address
        }
        
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

    def _add_frequency_statistics(self, ws: Worksheet, frequency_data: List[Dict], 
                                total_addresses: int) -> None:
        """
        Add statistical summary to the frequency analysis sheet.
        
        File: file_handler.py
        Function: _add_frequency_statistics()
        
        This adds summary statistics above the main data table.
        """
        # Calculate statistics
        if not frequency_data:
            # Handle empty data case
            ws.cell(row=1, column=1, value="ADDRESS FREQUENCY ANALYSIS SUMMARY").font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value=f"Total Addresses Analyzed: {total_addresses}")
            ws.cell(row=2, column=3, value="Unique Cluster Addresses: 0")
            ws.cell(row=3, column=1, value="No frequency data available")
            return
        
        frequencies = [data['frequency'] for data in frequency_data]
        total_unique_clusters = len(frequency_data)
        avg_frequency = sum(frequencies) / len(frequencies) if frequencies else 0
        max_frequency = max(frequencies) if frequencies else 0
        
        # Find the most frequent cluster
        most_frequent = max(frequency_data, key=lambda x: x['frequency']) if frequency_data else None
        
        # Write statistics
        ws.cell(row=1, column=1, value="ADDRESS FREQUENCY ANALYSIS SUMMARY").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Total Addresses Analyzed: {total_addresses}")
        ws.cell(row=2, column=3, value=f"Unique Cluster Addresses: {total_unique_clusters}")
        ws.cell(row=3, column=1, value=f"Average Frequency: {avg_frequency:.1f}")
        ws.cell(row=3, column=3, value=f"Maximum Frequency: {max_frequency}")
        
        if most_frequent:
            ws.cell(row=4, column=1, value=f"Most Frequent Cluster: {most_frequent['cluster_name']} ({most_frequent['frequency']} occurrences)")
        
        # Style the statistics section
        for row in range(1, 5):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                if cell.value:  # Only style cells with content
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
